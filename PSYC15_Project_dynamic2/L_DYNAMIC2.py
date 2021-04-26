#Emily Crowl
#Emotional Recognition: DYNAMIC 2
#PSYC15

#Collects user information:
#ID
#Runs 36 videos with each of the 6 basic emotions
#Users enter one word that describes which emotion they see
#program records response, marks right/wrong, and records time

from tkinter import ttk, Tk, Toplevel
from tkinter import *
import time
import random
import os
import os.path
import calcData
import response
import imageio
import threading
import moviepy.editor as mp
from openpyxl import *
from openpyxl.utils import get_column_letter
import dictionary_vid

#in case we are running this on a computer
#that doesn't have updated Python
try:
    from PIL import Image, ImageTk
except ModuleNotFoundError:
    import Image, ImageTk

#root window
root = Tk()

#initialize windows
#welcome window
welcome_w = Toplevel(root)
width_value = welcome_w.winfo_screenwidth()
height_value = welcome_w.winfo_screenheight()
welcome_w.configure(bg = 'white')
welcome_w.attributes('-fullscreen', True)
welcome_w.resizable(width = False, height = False)
welcome_w.title('Welcome')

#info window
info_w = Toplevel(root)
width_value = info_w.winfo_screenwidth()
height_value = info_w.winfo_screenheight()
info_w.configure(bg = 'white')
info_w.attributes('-fullscreen', True)
info_w.resizable(width = False, height = False)
info_w.title('User Information')

#instructions window
instruct_w = Toplevel(root)
width_value = instruct_w.winfo_screenwidth()
height_value = instruct_w.winfo_screenheight()
instruct_w.configure(bg = 'white')
instruct_w.attributes('-fullscreen', True)
instruct_w.resizable(width = False, height = False)
instruct_w.title('Instructions')

#start window; manip in recursive function
start_w = Toplevel(root)
width_value = start_w.winfo_screenwidth()
height_value = start_w.winfo_screenheight()
start_w.configure(bg = 'white')
start_w.attributes('-fullscreen', True)
start_w.resizable(width = False, height = False)
start_w.title('Start')

#label we are manipulating in recursive function
my_pic = ttk.Label(start_w, text = "Press \"Begin\" when you are ready.",\
                   background = 'white')
my_pic.config(font = ("Calibri", 18), justify = 'center')

my_pic.pack(side = "top", pady = (100, 0))

#videos conversion
##img1 = "vid1.mp4"
##img2 = "vid2.mp4"
##img3 = "vid3.mp4"
##img4 = "vid4.mp4"
##img5 = "vid5.mp4"
##img6 = "vid6.mp4"
##img7 = "vid7.mp4"
##img8 = "vid8.mp4"
##img9 = "vid9.mp4"
##img10 = "vid10.mp4"
##img11 = "vid11.mp4"
##img12 = "vid12.mp4"
##img13 = "vid13.mp4"
##img14 = "vid14.mp4"
##img15 = "vid15.mp4"
##img16 = "vid16.mp4"
##img17 = "vid17.mp4"
##img18 = "vid18.mp4"
##img19 = "vid19.mp4"
##img20 = "vid20.mp4"
##img21 = "vid21.mp4"
##img22 = "vid22.mp4"
##img23 = "vid23.mp4"
##img24 = "vid24.mp4"
##img25 = "vid25.mp4"
##img26 = "vid26.mp4"
##img27 = "vid27.mp4"
##img28 = "vid28.mp4"
##img29 = "vid29.mp4"
##img30 = "vid30.mp4"
##img31 = "vid31.mp4"
##img32 = "vid32.mp4"
##img33 = "vid33.mp4"
##img34 = "vid34.mp4"
##img35 = "vid35.mp4"
##img36 = "vid36.mp4"
##
##imgLAST = PhotoImage(file = "LAST.png")
##
##image_list1 = [img1, img2, img3, img4, img5, img6, img7, img8,\
##               img9, img10, img11, img12, img13, img14, img15,\
##               img16, img17, img18, img19, img20, img21, img22,\
##               img23, img24, img25, img26, img27, img28, img29,\
##               img30, img31, img32, img33, img34, img35, img36]
##
##
##
###one time resize code, comment out when done
##image_list = []
##index = 0
##count = 111
##
##for i in image_list1:
##    
##    count1 = str(count)
##    clip = mp.VideoFileClip(image_list1[index])
##    clip_resized = clip.resize(height = 700)
##    clip_resized.write_videofile("vid" + count1 + ".mp4")
##    imgnew  = "vid" + count1 + ".mp4"
##    index += 1
##    count += 1
##    image_list.append(imgnew)

#converted videos
img1 = "vid111.mp4"
img2 = "vid112.mp4"
img3 = "vid113.mp4"
img4 = "vid114.mp4"
img5 = "vid115.mp4"
img6 = "vid116.mp4"
img7 = "vid117.mp4"
img8 = "vid118.mp4"
img9 = "vid119.mp4"
img10 = "vid120.mp4"
img11 = "vid121.mp4"
img12 = "vid122.mp4"
img13 = "vid123.mp4"
img14 = "vid124.mp4"
img15 = "vid125.mp4"
img16 = "vid126.mp4"
img17 = "vid127.mp4"
img18 = "vid128.mp4"
img19 = "vid129.mp4"
img20 = "vid130.mp4"
img21 = "vid131.mp4"
img22 = "vid132.mp4"
img23 = "vid133.mp4"
img24 = "vid134.mp4"
img25 = "vid135.mp4"
img26 = "vid136.mp4"
img27 = "vid137.mp4"
img28 = "vid138.mp4"
img29 = "vid139.mp4"
img30 = "vid140.mp4"
img31 = "vid141.mp4"
img32 = "vid142.mp4"
img33 = "vid143.mp4"
img34 = "vid144.mp4"
img35 = "vid145.mp4"
img36 = "vid146.mp4"

imgLAST = PhotoImage(file = "LAST.png")

image_list = [img1, img2, img3, img4, img5, img6, img7, img8,\
              img9, img10, img11, img12, img13, img14, img15,\
              img16, img17, img18, img19, img20, img21, img22,\
              img23, img24, img25, img26, img27, img28, img29,\
              img30, img31, img32, img33, img34, img35, img36]

image_list_OG = [img1, img2, img3, img4, img5, img6, img7, img8,\
              img9, img10, img11, img12, img13, img14, img15,\
              img16, img17, img18, img19, img20, img21, img22,\
              img23, img24, img25, img26, img27, img28, img29,\
              img30, img31, img32, img33, img34, img35, img36]

#image randomization
random.shuffle(image_list)
trial_list = image_list

#hides windows on command
root.withdraw()
info_w.withdraw()
instruct_w.withdraw()
start_w.withdraw()

#functions to direct which window to open next
def goto_info():
    welcome_w.destroy()
    info_w.deiconify()

def goto_instructions():
    info_w.destroy()
    instruct_w.deiconify()

def goto_start():
    instruct_w.destroy()
    start_w.deiconify()

#functions for entries
def folder(studid, trial_list):
    name = userid.get()
    print(name, "Folder already exists, extracting information")
    user_file = open(name + '/' + 'User Information' + name + '.txt', 'r')
    info_list = []
    for i in user_file:
        info_list.append(i.strip('\n'))
    info_values(info_list, trial_list)
    user_file.close()

def info_values(info_list, trial_list):
    username = info_list[0]
    userage = info_list[1] #added age
    usergender = info_list[2] #added gender
    userpsyc = info_list[4]
    userprof = info_list[5]
    useronset = info_list[6]
    userdegree = info_list[7]
    usersign = info_list[8]
    userlang = info_list[9]
    userparents = info_list[10]

    #open or create workbook and sheets
    try:
        excel = load_workbook('C:/Users/emily/Desktop/em_recog_dynamic.xlsx')
        info_sheet = excel['INFO']
        data_sheet = excel.create_sheet('DATA', 1)
    except FileNotFoundError:
        excel = Workbook()
        info_sheet = excel.create_sheet('INFO', 0)
        data_sheet = excel.create_sheet('DATA', 1)
        
    insert_info(excel, info_sheet, data_sheet, username, userage, usergender, userid, userpsyc, userprof, useronset, userdegree, usersign, userlang, userparents, trial_list)

def mark_image(img_num, trial_list, user_emo, studid, excel, data_sheet):
    if img_num != 37:
        user_answer = user_emo.get()
        
        video = str(trial_list[img_num])
        values = dictionary_vid.getKey(video)
        
        count = 0
        for i in values:
            if user_answer.lower() == i:
                count += 1

            else:
                count += 0
                
        if count != 1:
            mark = "I/V"
            accuracy(studid, mark, img_num)
            data_sheet.cell(row = img_num + 2, column = 4).value = mark
            data_sheet.cell(row = img_num + 2, column = 5).value = values[0]
            excel.save('C:/Users/emily/Desktop/em_recog_dynamic.xlsx')
            
        else:
            mark = "C"
            accuracy(studid, mark, img_num)
            data_sheet.cell(row = img_num + 2, column = 4).value = mark
            excel.save('C:/Users/emily/Desktop/em_recog_dynamic.xlsx')

def accuracy(studid, mark, img_num):
    studid = userid.get()
    
    save_to_D2 = 'C:/Users/emily/Desktop/PSYC15_Project_dynamic2//' + studid
    completeName_D2 = os.path.join(save_to_D2, 'accuracy_dynamic' + studid + '.txt')
    
    if not os.path.exists(save_to_D2):
        outfile_D2 = open(completeName_D2, 'w')
        outfile_D2.write(str(img_num+1) + ') ' + str(mark) + '\n')
        outfile_D2.close()
    else:
        outfile_D2 = open(completeName_D2, 'a')
        outfile_D2.write(str(img_num + 1) + ') ' + str(mark) + '\n')
        outfile_D2.close()

#img order    
def img_order(studid, trial_list):
    studid = userid.get()
    save_path = 'C:/Users/emily/Desktop/PSYC15_Project_dynamic2//' + studid
    completeName = os.path.join(save_path, 'Image Order' + studid + '.txt')
    outfile = open(completeName, 'w')
    outfile.write(str('Random Image Order for ' + studid + '\n'))
    outfile.close
    for i in trial_list:
        save_path = 'C:/Users/emily/Desktop/PSYC15_Project_dynamic2//' + studid
        completeName = os.path.join(save_path, 'Image Order' + studid + '.txt')
        outfile = open(completeName, 'a')
        outfile.write(str(i) + '\n')
    outfile.close()
    
#functions for buttons
def writeclick(studid, click):
    studid = userid.get()
    click = user_emo.get()
    outfile = open('Raw Data.txt', 'a')
    outfile.write(str(click) + '\n')
    outfile.close()

def appendclick(studid, click):
    studid = userid.get()
    click = user_emo.get()
    outfile = open('Raw Data.txt', 'a') 
    outfile.write(str(click) + '\n')
    outfile.close()

#functions to write in times marked by button clicks
#function to capture clicks
def timeclickStart(event):
    studid = userid.get()
    start = time.time()
    outfile = open('Raw Data.txt', 'w')
    outfile.write(str(studid) + '\n')
    outfile.write(str(start) + '\n') 
    outfile.close()

def timeclick(*args):
    studid = userid.get()
    initial = time.time()
    outfile = open('Raw Data.txt', 'a')
    outfile.write(str(initial) + '\n')
    outfile.close()
    
def timeclick1(*args):
    studid = userid.get()
    post = time.time()
    outfile = open('Raw Data.txt', 'a')
    outfile.write(str(post) + '\n')
    outfile.close()

#function to invoke the submit button when enter key or mouse click is pressed
def buttoncall(*args):
    submit.invoke()

#function to clear entry box
def clear(*args):
    enter.delete(first = 0, last = 25)


#pack submit and enter labels
def entrypack():
    submit.configure(state=NORMAL)
    submit.bind("<Button-1>", buttoncall)
    submit.pack(side = "bottom", pady = (10, 150))
    
    enter.configure(state=NORMAL)
    enter.bind("<Return>", buttoncall)
    enter.focus_set()
    enter.pack(side = 'bottom')

    em_list.pack()
    em_list.place(x = 900, y = 550)
    
#functions for info next buttons
def buttonpressed1(*args):
    next2.pack(side = "bottom", pady = (0, 200))

def buttonpressed2(*args):
    next3.pack(side = 'bottom', pady = (0, 100))

#function for goodbye window
def goodbye():
    root.title("Thank you")
    global button_exit
    button_exit.pack_forget()

    end_label = ttk.Label(start_w, text = "\n\n\nThank you for your participation in this task.", \
                          background = "white")
    end_label.config(font = ("Calibri", 15), justify = 'center')
    
    end_label.pack(side = "top", pady = (20, 0))

def insert_info(excel, info_sheet, data_sheet, username, userage, usergender, userid, userpsyc, userprof, useronset, userdegree, usersign, userlang, userparents, trial_list):
    #info sheet
    info_sheet.column_dimensions['A'].width = 25
    info_sheet.column_dimensions['B'].width = 7 #age column
    info_sheet.column_dimensions['C'].width = 10 # gender column
    info_sheet.column_dimensions['D'].width = 15
    info_sheet.column_dimensions['E'].width = 15
    info_sheet.column_dimensions['F'].width = 15
    info_sheet.column_dimensions['G'].width = 20
    info_sheet.column_dimensions['H'].width = 20
    info_sheet.column_dimensions['I'].width = 20
    info_sheet.column_dimensions['J'].width = 15
    info_sheet.column_dimensions['K'].width = 15
    info_sheet.column_dimensions['L'].width = 15

    #data sheet
    data_sheet.column_dimensions['A'].width = 15
    data_sheet.column_dimensions['B'].width = 15
    data_sheet.column_dimensions['C'].width = 15
    data_sheet.column_dimensions['D'].width = 20

    #info labels
    info_sheet.cell(row = 1, column = 1).value = "Name"
    info_sheet.cell(row = 1, column = 2).value = "Age"
    info_sheet.cell(row = 1, column = 3).value = "Gender" #NEW
    info_sheet.cell(row = 1, column = 4).value = "Student ID"
    info_sheet.cell(row = 1, column = 5).value = "Class"
    info_sheet.cell(row = 1, column = 6).value = "Professor"
    info_sheet.cell(row = 1, column = 7).value = "Onset hearing loss"
    info_sheet.cell(row = 1, column = 8).value = "Severity"
    info_sheet.cell(row = 1, column = 9).value = "Signing/non-signing"
    info_sheet.cell(row = 1, column = 10).value = "Language learning"
    info_sheet.cell(row = 1, column = 11).value = "Parents"
    info_sheet.cell(row = 1, column = 12).value = "IC"

    #data labels
    data_sheet.cell(row = 1, column = 1).value = "Image Order"
    data_sheet.cell(row = 1, column = 2).value = "Time"
    data_sheet.cell(row = 1, column = 3).value = "Response"
    data_sheet.cell(row = 1, column = 4).value = "Correct/Incorrect"
    data_sheet.cell(row = 1, column = 5).value = "Answer"

    #info 
    #get method returns current text as string, placing in specific location
    current_row = info_sheet.max_row #THE ONE LINE OF CODE THAT SEPARATED GOOD AND EVIL

    info_sheet.cell(row = current_row + 1, column = 1).value = username
    info_sheet.cell(row = current_row + 1, column = 2).value = userage
    info_sheet.cell(row = current_row + 1, column = 3).value = usergender #gender
    info_sheet.cell(row = current_row + 1, column = 4).value = userid.get()
    info_sheet.cell(row = current_row + 1, column = 5).value = userpsyc
    info_sheet.cell(row = current_row + 1, column = 6).value = userprof
    info_sheet.cell(row = current_row + 1, column = 7).value = useronset
    info_sheet.cell(row = current_row + 1, column = 8).value = userdegree
    info_sheet.cell(row = current_row + 1, column = 9).value = usersign
    info_sheet.cell(row = current_row + 1, column = 10).value = userlang
    info_sheet.cell(row = current_row + 1, column = 11).value = userparents
    info_sheet.cell(row = current_row + 1, column = 12).value = None
    

    #image order
    order_row = 2
    for i in trial_list:
        j = str(i)
        data_sheet.cell(row = order_row, column = 1).value = j
        order_row += 1
        
    #save file
    data_sheet.title = 'DATA'   
    excel.save('C:/Users/emily/Desktop/em_recog_dynamic.xlsx')
    
def insert_data(excel, data_sheet, user_emo, img_num):
    data_sheet.cell(row = img_num, column = 3).value = user_emo.get()
    
    excel.save('C:/Users/emily/Desktop/em_recog_dynamic.xlsx')

def collate(image_list_OG):
    name = userid.get()
    print(name, "Folder already exists, extracting information")
    user_file = open(name + '/' + 'User Information' + name + '.txt', 'r')
    info_list = []
    for i in user_file:
        info_list.append(i.strip('\n'))

    username = info_list[0]
    userage = info_list[1]
    usergender = info_list[2] #NEW
    userpsyc = info_list[4]
    userprof = info_list[5]
    useronset = info_list[6]
    userdegree = info_list[7]
    usersign = info_list[8]
    userlang = info_list[9]
    userparents = info_list[10]
    
    try:
        excel = load_workbook('C:/Users/emily/Desktop/em_recog_dynamic_cleaned.xlsx')
        sort_sheet = excel['CLEANED']

        current_row = sort_sheet.max_row
    
        sort_sheet.cell(row = current_row + 1, column = 1).value = current_row 
        sort_sheet.cell(row = current_row + 1, column = 2).value = userage
        sort_sheet.cell(row = current_row + 1, column = 3).value = usergender #NEW
        sort_sheet.cell(row = current_row + 1, column = 4).value = useronset
        sort_sheet.cell(row = current_row + 1, column = 5).value = userdegree
        sort_sheet.cell(row = current_row + 1, column = 6).value = usersign
        sort_sheet.cell(row = current_row + 1, column = 7).value = userlang
        sort_sheet.cell(row = current_row + 1, column = 8).value = userparents

        excel.save(filename = 'C:/Users/emily/Desktop/em_recog_dynamic_cleaned.xlsx')

    except FileNotFoundError:
        excel = Workbook()
        sort_sheet = excel.create_sheet("CLEANED", 0)

        data_column = 4
        while data_column < 9:
            letter_col = get_column_letter(data_column)
            sort_sheet.column_dimensions[letter_col].width = 20
            data_column += 1

        order_column = 9
        time_column = 45
        order_row = 1
        for i in image_list_OG:
            j = str(i)
            values = dictionary_vid.getKey(j)
            letter = get_column_letter(order_column)
            letter_time = get_column_letter(time_column)
            sort_sheet.column_dimensions[letter].width = 20
            sort_sheet.column_dimensions[letter_time].width = 20
            sort_sheet.cell(row = order_row, column = order_column).value = j + " /" + values[0]
            sort_sheet.cell(row = order_row, column = time_column).value = j + " /" + values[0]
            order_column += 1
            time_column += 1

        sort_sheet.cell(row = 1, column = 2).value = "Age"
        sort_sheet.cell(row = 1, column = 3).value = "Gender" #NEW
        sort_sheet.cell(row = 1, column = 4).value = "Onset of Hearing Loss"
        sort_sheet.cell(row = 1, column = 5).value = "Severity"
        sort_sheet.cell(row = 1, column = 6).value = "Signing/Non-signing"
        sort_sheet.cell(row = 1, column = 7).value = "Language"
        sort_sheet.cell(row = 1, column = 8).value = "Parents"

        current_row = sort_sheet.max_row
    
        sort_sheet.cell(row = current_row + 1, column = 1).value = current_row 
        sort_sheet.cell(row = current_row + 1, column = 2).value = userage
        sort_sheet.cell(row = current_row + 1, column = 3).value = usergender #NEW
        sort_sheet.cell(row = current_row + 1, column = 4).value = useronset
        sort_sheet.cell(row = current_row + 1, column = 5).value = userdegree
        sort_sheet.cell(row = current_row + 1, column = 6).value = usersign
        sort_sheet.cell(row = current_row + 1, column = 7).value = userlang
        sort_sheet.cell(row = current_row + 1, column = 8).value = userparents

        excel.save(filename = 'C:/Users/emily/Desktop/em_recog_dynamic_cleaned.xlsx')
        

def readData():
    studid = userid.get()
    file = open("C:/Users/emily/Desktop/PSYC15_Project_dynamic2//" + "/" + studid + "/" + "accuracy_dynamic" + studid + ".txt", "r")
    listRead = file.readlines()
    acc = []
    count = 1

    for i in listRead:
        j = i.replace(str(count),"")
        k = j.replace(")", '')
        l = k.replace("\n", '')
        acc.append(l)
        count += 1
    
    studid = userid.get()
    file1 = open("C:/Users/emily/Desktop/PSYC15_Project_dynamic2//" + "/" + studid + "/" + "Image Order" + studid + ".txt", "r")
    listRead1 = file1.readlines()
    
    img = []
    del listRead1[0]
    for i in listRead1:
        j = i.replace("\n", '')
        k = j[3:6]
        img.append(int(k))

    file2 = open("C:/Users/emily/Desktop/PSYC15_Project_dynamic2//" + "/" + studid + "/" + "Time" + studid + ".txt", "r")
    listRead_time = file2.readlines()
    time = []

    del listRead_time[0]
    del listRead_time[0]
    for i in listRead_time:
        j = i.replace("\n", "")
        k = j[3:13]
        time.append(k)
    #print(time)
    
    index = 0
    time_dict= {}
    new_dict = {}
    for i in range(0, len(img)):
        key = img[i]
        value = acc[index]
        time_value = time[index]
        index += 1
        new_dict[key] = value
        time_dict[key] = time_value

    excel = load_workbook('C:/Users/emily/Desktop/em_recog_dynamic_cleaned.xlsx')
    sort_sheet = excel['CLEANED']

    current_row = sort_sheet.max_row
    order_column = 9
    time_column = 45

    x = sorted(new_dict)
    y = sorted(time_dict)

    for i in y:
        j = new_dict[i]
        k = time_dict[i]
        sort_sheet.cell(row = current_row, column = order_column).value = j
        sort_sheet.cell(row = current_row, column = time_column).value = k
        order_column += 1
        time_column += 1
    
    excel.save(filename = 'C:/Users/emily/Desktop/em_recog_dynamic_cleaned.xlsx')

#recursive function to advance the imgs
def forward(img_num):
    global my_pic
    global start
    global submit
    global button_exit
    global studid
    global enter
    global em_list
    start_w.title(' ')
    excel = load_workbook('C:/Users/emily/Desktop/em_recog_dynamic.xlsx')
    data_sheet = excel['DATA']

#similar to windows
    my_pic.pack_forget()
    start.pack_forget()
    submit.pack_forget()
    enter.pack_forget()
    em_list.pack_forget()
    em_list.place_forget()

    enter.delete(0,END)
    submit.configure(state = DISABLED)
    enter.configure(state = DISABLED)

#trivial statement
    if img_num == 37:
        em_list.pack_forget()
        em_list.place_forget()
        my_pic = Label(start_w, image = imgLAST, bg = "white")
        my_pic.pack(side = "top", pady = (75,0))

        button_exit = Button(start_w, text = "Quit", command = lambda:[calcData.calcData(), response.organize(), goodbye(),\
                                                                       collate(image_list_OG), readData()]) #added functions here
        
        button_exit.pack(side = "bottom", pady = (0, 300))

        goodbye_label = ttk.Label(start, text = "Please press quit.", background = "white")
        goodbye_label.pack(side = "bottom", pady = (0,0))

#advancing through imgs
    elif img_num == 1:
        def stream(label):
            video = imageio.get_reader(trial_list[img_num - 1])
            
            for vid in video.iter_data():
                frame_vid = ImageTk.PhotoImage(Image.fromarray(vid))
                label.config(image = frame_vid)
                label.image = frame_vid

        thread = threading.Thread(target = stream, args = (my_pic,))
        thread.start() 

        my_pic.pack(side = "top")
        
        submit = Button(start_w, text = "Submit", state = DISABLED, command = lambda:[timeclick1(), appendclick(studid, click = user_emo.get()),\
                                                                 insert_data(excel, data_sheet, user_emo, img_num + 1),\
                                                                 mark_image(img_num - 1, trial_list, user_emo, studid, excel, data_sheet),\
                                                                 forward(img_num + 1), clear()])
        
        em_list = Label(start_w, text = "HAPPY\n\nSAD\n\nANGRY\n\nSURPRISE\n\nDISGUST\n\nAFRAID", background = 'white')

        enter = ttk.Entry(start_w, state = DISABLED, textvariable = user_emo)

        submit.after(5500, entrypack)
    
    elif img_num % 2 == 0:
        def stream(label):
            video = imageio.get_reader(trial_list[img_num - 1])
            
            for vid in video.iter_data():
                frame_vid = ImageTk.PhotoImage(Image.fromarray(vid))
                label.config(image = frame_vid)
                label.image = frame_vid

        thread = threading.Thread(target = stream, args = (my_pic,))
        thread.start()

        my_pic.pack(side = "top")
        
        submit = Button(start_w, text = "Submit", state = DISABLED, command = lambda:[timeclick(), appendclick(studid, click = user_emo.get()),\
                                                                 insert_data(excel, data_sheet, user_emo, img_num + 1),\
                                                                 mark_image(img_num - 1, trial_list, user_emo, studid, excel, data_sheet),\
                                                                 forward(img_num + 1), clear()])

        enter = ttk.Entry(start_w, state = DISABLED, textvariable = user_emo)
        
        em_list = Label(start_w, text = "HAPPY\n\nSAD\n\nANGRY\n\nSURPRISE\n\nDISGUST\n\nAFRAID", background = 'white')

        submit.after(5500, entrypack)
        
    elif img_num != 1 and img_num % 2 != 0 and img_num != 37:
        def stream(label):
            video = imageio.get_reader(trial_list[img_num - 1])
            
            for vid in video.iter_data():
                frame_vid = ImageTk.PhotoImage(Image.fromarray(vid))
                label.config(image = frame_vid)
                label.image = frame_vid

        thread = threading.Thread(target = stream, args = (my_pic,))
        thread.start()

        my_pic.pack(side = "top")
        
        submit = Button(start_w, text = "Submit", state=DISABLED, command = lambda:[timeclick1(), appendclick(studid, click = user_emo.get()),\
                                                                 insert_data(excel, data_sheet, user_emo, img_num + 1),\
                                                                 mark_image(img_num - 1, trial_list, user_emo, studid, excel, data_sheet),\
                                                                 forward(img_num + 1), clear()])

        enter = ttk.Entry(start_w, state = DISABLED, textvariable = user_emo)

        em_list = Label(start_w, text = "HAPPY\n\nSAD\n\nANGRY\n\nSURPRISE\n\nDISGUST\n\nAFRAID", background = 'white')

        submit.after(5500, entrypack)
 
#main
def windows():
#welcome page layout
    wel_label = ttk.Label(welcome_w, text = \
                          """\n\nWelcome!\n\n\n\nThe following page will ask you for the following information:
\nstudent ID
\nPlease proceed to the next page.""",\
                          background = 'white')
    wel_label.config(font = ("Calibri", 18), justify = 'center')
    wel_label.pack(side = "top", pady = (50, 0))

    #button layout
    next1 = ttk.Button(welcome_w, text = 'Next',\
                       command = lambda:[goto_info()])
    next1.bind("<Button-1>")
    next1.pack(side = "bottom", pady = (0, 250))

#info page layout
    info_label = ttk.Label(info_w, text = "Please enter your student ID:", background = "white")
    info_label.config(font = ("Calibri", 12), justify = 'right')
    info_label.pack(side = "top", pady = (150, 0))

    #student id
    id_label = ttk.Label(info_w, text = "IVC or Saddleback Student ID: ", background = 'white')
    id_label.pack()
    id_label.place(x = 494, y = 250)
    
    id_entry = ttk.Entry(info_w, textvariable = userid)
    id_entry.pack()
    id_entry.place(x = 700, y = 250)

    #advance
    confirm = ttk.Button(info_w, text = 'The above information is correct.',\
                         command = lambda:[folder(studid, trial_list), img_order(studid, trial_list), buttonpressed1()])
    confirm.bind("<Button-1>")
    confirm.pack()
    confirm.place(x = 590, y = 450)
    
    next2 = ttk.Button(info_w, text = 'Next',\
                       command = lambda:[ goto_instructions()])
    next2.bind("<Button-1>")

#instructions page layout
    instruct = ttk.Label(instruct_w, text = \
                   """Participation in this task is voluntary.\nYou may choose to opt out at any time and your data will not be recorded.
                    \nInstructions are below.
                    \nInstructions:\nFor each video, enter exactly one word to describe the emotion that you see as quickly and as accurately as possible,
                    then click the "Enter" key on the keyboard or the "Submit" button.
                    \n***Let the video play in its entirety***
                    \nWhen you are done, please press the \"Quit\" button and leave the window up.
                    \nPlease ensure you have read all the instructions and ask the researcher if you have any questions.
                    \n\nPress the button below to confirm you have read the instructions and press \"Next\".""",\
                   background = 'white')
    instruct.config(font = ("Calibri", 18), justify = 'center')

    instruct.pack(side = "top", pady = (100, 0))

    confirm_dir_read = ttk.Button(instruct_w, text = 'I have read and I understand the above instructions.',\
                             command = lambda:[buttonpressed2()])
    confirm_dir_read.bind("<Button-1")
    confirm_dir_read.place(x = 545, y = 575)
    confirm_dir_read.pack()

    next3 = ttk.Button(instruct_w, text = 'Next',\
                       command = lambda:[goto_start()])
    next3.bind("<Button-1>")

#declared variables for global/function use  
#info
#string variables
userid = StringVar()

#user id
studid = userid.get()

#user entry
user_emo = StringVar()
enter = ttk.Entry(start_w, textvariable = user_emo)

#buttons
#info next
next2 = ttk.Button(info_w, text = 'Next',\
                       command = lambda:[goto_instructions()])

next3 = ttk.Button(instruct_w, text = 'Next',\
                       command = lambda:[goto_start()])

#start
start = ttk.Button(start_w, text = 'Begin', command = lambda:[forward(1)])
start.bind("<Button-1>", timeclickStart)
start.pack(side = 'bottom', pady = (0, 500))

#enter key
submit = Button(start_w, text = "Submit", command = lambda:[timeclick(), appendclick(studid, click = user_emo.get()),\
                                                                 insert_data(excel, data_sheet, user_emo, img_num + 1), forward(img_num + 1), clear()])

em_list = Label(start_w, text = "HAPPY\n\nSAD\n\nANGRY\n\nSURPRISE\n\nDISGUST\n\nAFRAID", background = 'white')

#quit
button_exit = Button(start_w, text = "Quit", command = lambda:[calcData.calcData(), response.organize(), goodbye()])

if __name__=="__main__":
    windows()
    
root.mainloop()
