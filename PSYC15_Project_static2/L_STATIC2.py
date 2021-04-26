#Emily Crowl
#Emotional Recognition
#PSYC15

#Collects user information:
#ID
#Runs 60 static images with each of the 6 basic emotions
#Users enter one word that describes which emotion they see
#program records response, marks right/wrong, and records time

from tkinter import ttk, Tk, Toplevel
from tkinter import *
import time
import random
import os
import os.path
import calcData
import response
from openpyxl import *
from openpyxl.utils import get_column_letter
import dictionary_img

#in case we are running this on a computer
#that doesn't have updated Python
try:
    from PIL import Image
except ModuleNotFoundError:
    import Image

#root window
root = Tk()

#initialize windows
#welcome window
welcome_w = Toplevel(root)
width_value = welcome_w.winfo_screenwidth()
height_value = welcome_w.winfo_screenheight()
welcome_w.configure(bg = 'white')
welcome_w.attributes('-fullscreen', True)
welcome_w.resizable(width = False, height = False)
welcome_w.title('Welcome')

#info window
info_w = Toplevel(root)
width_value = info_w.winfo_screenwidth()
height_value = info_w.winfo_screenheight()
info_w.configure(bg = 'white')
info_w.attributes('-fullscreen', True)
info_w.resizable(width = False, height = False)
info_w.title('User Information')

#instructions window
instruct_w = Toplevel(root)
width_value = instruct_w.winfo_screenwidth()
height_value = instruct_w.winfo_screenheight()
instruct_w.configure(bg = 'white')
instruct_w.attributes('-fullscreen', True)
instruct_w.resizable(width = False, height = False)
instruct_w.title('Instructions')

#start window; manip in recursive function
start_w = Toplevel(root)
width_value = start_w.winfo_screenwidth()
height_value = start_w.winfo_screenheight()
start_w.configure(bg = 'white')
start_w.attributes('-fullscreen', True)
start_w.resizable(width = False, height = False)
start_w.title('Instructions')

#label we are manipulating in recursive function
my_pic = ttk.Label(start_w, text = "Press \"Begin\" when you are ready.""",\
                   background = 'white')
my_pic.config(font = ("Calibri", 18), justify = 'center')

my_pic.pack(side = "top", pady = (100, 0))

#imgs
img1 = PhotoImage(file = "pic1.png")
img2 = PhotoImage(file = "pic2.png")
img3 = PhotoImage(file = "pic3.png")
img4 = PhotoImage(file = "pic4.png")
img5 = PhotoImage(file = "pic5.png")
img6 = PhotoImage(file = "pic6.png")
img7 = PhotoImage(file = "pic7.png")
img8 = PhotoImage(file = "pic8.png")
img9 = PhotoImage(file = "pic9.png")
img10 = PhotoImage(file = "pic10.png")
img11 = PhotoImage(file = "pic11.png")
img12 = PhotoImage(file = "pic12.png")
img13 = PhotoImage(file = "pic13.png")
img14 = PhotoImage(file = "pic14.png")
img15 = PhotoImage(file = "pic15.png")
img16 = PhotoImage(file = "pic16.png")
img17 = PhotoImage(file = "pic17.png")
img18 = PhotoImage(file = "pic18.png")
img19 = PhotoImage(file = "pic19.png")
img20 = PhotoImage(file = "pic20.png")
img21 = PhotoImage(file = "pic21.png")
img22 = PhotoImage(file = "pic22.png")
img23 = PhotoImage(file = "pic23.png")
img24 = PhotoImage(file = "pic24.png")
img25 = PhotoImage(file = "pic25.png")
img26 = PhotoImage(file = "pic26.png")
img27 = PhotoImage(file = "pic27.png")
img28 = PhotoImage(file = "pic28.png")
img29 = PhotoImage(file = "pic29.png")
img30 = PhotoImage(file = "pic30.png")
img31 = PhotoImage(file = "pic31.png")
img32 = PhotoImage(file = "pic32.png")
img33 = PhotoImage(file = "pic33.png")
img34 = PhotoImage(file = "pic34.png")
img35 = PhotoImage(file = "pic35.png")
img36 = PhotoImage(file = "pic36.png")
img37 = PhotoImage(file = "pic37.png")
img38 = PhotoImage(file = "pic38.png")
img39 = PhotoImage(file = "pic39.png")
img40 = PhotoImage(file = "pic40.png")
img41 = PhotoImage(file = "pic41.png")
img42 = PhotoImage(file = "pic42.png")
img43 = PhotoImage(file = "pic43.png")
img44 = PhotoImage(file = "pic44.png")
img45 = PhotoImage(file = "pic45.png")
img46 = PhotoImage(file = "pic46.png")
img47 = PhotoImage(file = "pic47.png")
img48 = PhotoImage(file = "pic48.png")
img49 = PhotoImage(file = "pic49.png")
img50 = PhotoImage(file = "pic50.png")
img51 = PhotoImage(file = "pic51.png")
img52 = PhotoImage(file = "pic52.png")
img53 = PhotoImage(file = "pic53.png")
img54 = PhotoImage(file = "pic54.png")
img55 = PhotoImage(file = "pic55.png")
img56 = PhotoImage(file = "pic56.png")
img57 = PhotoImage(file = "pic57.png")
img58 = PhotoImage(file = "pic58.png")
img59 = PhotoImage(file = "pic59.png")
img60 = PhotoImage(file = "pic60.png")

imgLAST = PhotoImage(file = "LAST.png")

image_list = [img1, img2, img3, img4, img5, img6, img7, img8, img9, img10,\
              img11, img12, img13, img14, img15, img16, img17, img18,\
              img19, img20, img21, img22, img23, img24, img25, img26,\
              img27, img28, img29, img30, img31, img32, img33, img34,\
              img35, img36, img37, img38, img39, img40, img41, img42,\
              img43, img44, img45, img46, img47, img48, img49, img50,\
              img51, img52, img53, img54, img55, img56, img57, img58,\
              img59, img60]

image_list_OG = [img1, img7, img13, img19, img25, img31, img37, img43, img49, img55,\
                 img2, img8, img14, img20, img26, img32, img38, img44, img50, img56,\
                 img3, img9, img15, img21, img27, img33, img39, img45, img51, img57,\
                 img4, img10, img16, img22, img28, img34, img40, img46, img52, img58,\
                 img5, img11, img17, img23, img29, img35, img41, img47, img53, img59,\
                 img6, img12, img18, img24, img30, img36, img42, img48, img54, img60]

#image randomization
random.shuffle(image_list)
trial_list = image_list

#hides windows on command
root.withdraw()
info_w.withdraw()
instruct_w.withdraw()
start_w.withdraw()

#functions to direct which window to open next
def goto_info():
    welcome_w.destroy()
    info_w.deiconify()

def goto_instructions():
    info_w.destroy()
    instruct_w.deiconify()
    
def goto_start():
    instruct_w.destroy()
    start_w.deiconify()

#functions for entries
def folder(studid, trial_list):
    name = userid.get()
    print(name, "Folder already exists, extracting information")
    user_file = open(name + '/' + 'User Information' + name + '.txt', 'r')
    info_list = []
    for i in user_file:
        info_list.append(i.strip('\n'))
    info_values(info_list, trial_list)

def info_values(info_list, trial_list):
    username = info_list[0]
    userage = info_list[1] 
    usergender = info_list[2]
    userpsyc = info_list[4]
    userprof = info_list[5]
    useronset = info_list[6]
    userdegree = info_list[7]
    usersign = info_list[8]
    userlang = info_list[9]
    userparents = info_list[10] #open or create workbook and sheets
    try:
        excel = load_workbook('C:/Users/emily/Desktop/em_recog_static.xlsx')
        info_sheet = excel['INFO']
        data_sheet = excel.create_sheet('DATA', 1)
    except FileNotFoundError:
        excel = Workbook()
        info_sheet = excel.create_sheet('INFO', 0)
        data_sheet = excel.create_sheet('DATA', 1)
        
    insert_info(excel, info_sheet, data_sheet, username, userage, usergender, userid, userpsyc, userprof, useronset, userdegree, usersign, userlang, userparents, trial_list)

def mark_image(img_num, trial_list, user_emo, studid, excel, data_sheet):
    if img_num != 61:
        user_answer = user_emo.get()
        
        pyImage = str(trial_list[img_num])
        values = dictionary_img.getKey(pyImage)
        
        count = 0
        for i in values:
            if user_answer.lower() == i:
                count += 1

            else:
                count += 0
                
        if count != 1:
            mark = "I/V"
            accuracy(studid, mark, img_num)
            data_sheet.cell(row = img_num + 2, column = 4).value = mark
            data_sheet.cell(row = img_num + 2, column = 5).value = values[0]
            excel.save('C:/Users/emily/Desktop/em_recog_static.xlsx')
            
        else:
            mark = "C"
            accuracy(studid, mark, img_num)
            data_sheet.cell(row = img_num + 2, column = 4).value = mark
            excel.save('C:/Users/emily/Desktop/em_recog_static.xlsx')
            
def accuracy(studid, mark, img_num):
    studid = userid.get()

    save_to_S2 = 'C:/Users/emily/Desktop/PSYC15_Project_static2//' + studid
    completeName_S2 = os.path.join(save_to_S2, 'accuracy_static' + studid + '.txt')
    
    if not os.path.exists(save_to_S2):
        outfile_S2 = open(completeName_S2, 'w')
        outfile_S2.write(str(img_num+1) + ') ' + str(mark) + '\n')
        outfile_S2.close()
    else:
        outfile_S2 = open(completeName_S2, 'a')
        outfile_S2.write(str(img_num + 1) + ') ' + str(mark) + '\n')
        outfile_S2.close()

#img order    
def img_order(studid, trial_list):
    studid = userid.get()
    save_path = 'C:/Users/emily/Desktop/PSYC15_Project_static2//' + studid
    completeName = os.path.join(save_path, 'Image Order' + studid + '.txt')
    outfile = open(completeName, 'w')
    outfile.write(str('Random Image Order for ' + studid + '\n'))
    outfile.close
    for i in trial_list:
        save_path = 'C:/Users/emily/Desktop/PSYC15_Project_static2//' + studid
        completeName = os.path.join(save_path, 'Image Order' + studid + '.txt')
        outfile = open(completeName, 'a')
        outfile.write(str(i) + '\n')
    outfile.close()    

#functions for buttons
def writeclick(studid, click):
    studid = userid.get()
    click = user_emo.get()
    outfile = open('Raw Data.txt', 'a')
    outfile.write(str(click) + '\n')
    outfile.close()

def appendclick(studid, click):
    studid = userid.get()
    click = user_emo.get()
    outfile = open('Raw Data.txt', 'a') 
    outfile.write(str(click) + '\n')
    outfile.close()

#functions to write in times marked by button clicks
#function to capture clicks
def timeclickStart(event):
    studid = userid.get()
    start = time.time()
    outfile = open('Raw Data.txt', 'w')
    outfile.write(str(studid) + '\n')
    outfile.write(str(start) + '\n') 
    outfile.close()

def timeclick(*args):
    studid = userid.get()
    initial = time.time()
    outfile = open('Raw Data.txt', 'a')
    outfile.write(str(initial) + '\n')
    outfile.close()
    
def timeclick1(*args):
    studid = userid.get()
    post = time.time()
    outfile = open('Raw Data.txt', 'a')
    outfile.write(str(post) + '\n')
    outfile.close()

#function to invoke the submit button when enter key or mouse click is pressed
def buttoncall(*args):
    submit.invoke()

#function to clear entry box
def clear(*args):
    enter.delete(first = 0, last = 25)

#functions for info next buttons
def buttonpressed1(*args):
    next2.pack(side = "bottom", pady = (0, 200))

def buttonpressed2(*args):
    next3.pack(side =  'bottom', pady = (0, 100))

#function for goodbye window
def goodbye():
    root.title("Thank you")
    global button_exit
    button_exit.pack_forget()

    end_label = ttk.Label(start_w, text = "\n\n\nThank you for your participation in this task.", \
                          background = "white")
    end_label.config(font = ("Calibri", 15), justify = 'center')
    
    end_label.pack(side = "top", pady = (20, 0))

def insert_info(excel, info_sheet, data_sheet, username, userage, usergender, userid, userpsyc, userprof, useronset, userdegree, usersign, userlang, userparents, trial_list):
    #info sheet
    info_sheet.column_dimensions['A'].width = 25
    info_sheet.column_dimensions['B'].width = 7 
    info_sheet.column_dimensions['C'].width = 10
    info_sheet.column_dimensions['D'].width = 15
    info_sheet.column_dimensions['E'].width = 15
    info_sheet.column_dimensions['F'].width = 15
    info_sheet.column_dimensions['G'].width = 20
    info_sheet.column_dimensions['H'].width = 20
    info_sheet.column_dimensions['I'].width = 20
    info_sheet.column_dimensions['J'].width = 15
    info_sheet.column_dimensions['K'].width = 15
    info_sheet.column_dimensions['L'].width = 15

    #data sheet
    data_sheet.column_dimensions['A'].width = 15
    data_sheet.column_dimensions['B'].width = 15
    data_sheet.column_dimensions['C'].width = 15
    data_sheet.column_dimensions['D'].width = 20

    #info labels
    info_sheet.cell(row = 1, column = 1).value = "Name"
    info_sheet.cell(row = 1, column = 2).value = "Age"
    info_sheet.cell(row = 1, column = 3).value = "Gender"
    info_sheet.cell(row = 1, column = 4).value = "Student ID"
    info_sheet.cell(row = 1, column = 5).value = "Class"
    info_sheet.cell(row = 1, column = 6).value = "Professor"
    info_sheet.cell(row = 1, column = 7).value = "Onset hearing loss"
    info_sheet.cell(row = 1, column = 8).value = "Severity"
    info_sheet.cell(row = 1, column = 9).value = "Signing/non-signing"
    info_sheet.cell(row = 1, column = 10).value = "Language learning"
    info_sheet.cell(row = 1, column = 11).value = "Parents"
    info_sheet.cell(row = 1, column = 12).value = "IC"

    #data labels
    data_sheet.cell(row = 1, column = 1).value = "Image Order"
    data_sheet.cell(row = 1, column = 2).value = "Time"
    data_sheet.cell(row = 1, column = 3).value = "Response"
    data_sheet.cell(row = 1, column = 4).value = "Correct/Incorrect"
    data_sheet.cell(row = 1, column = 5).value = "Answer"

    #info 
    #get method returns current text as string, placing in specific location
    current_row = info_sheet.max_row #THE ONE LINE OF CODE THAT SEPARATED GOOD AND EVIL
    
    info_sheet.cell(row = current_row + 1, column = 1).value = username
    info_sheet.cell(row = current_row + 1, column = 2).value = userage
    info_sheet.cell(row = current_row + 1, column = 3).value = usergender #gender
    info_sheet.cell(row = current_row + 1, column = 4).value = userid.get()
    info_sheet.cell(row = current_row + 1, column = 5).value = userpsyc
    info_sheet.cell(row = current_row + 1, column = 6).value = userprof
    info_sheet.cell(row = current_row + 1, column = 7).value = useronset
    info_sheet.cell(row = current_row + 1, column = 8).value = userdegree
    info_sheet.cell(row = current_row + 1, column = 9).value = usersign
    info_sheet.cell(row = current_row + 1, column = 10).value = userlang
    info_sheet.cell(row = current_row + 1, column = 11).value = userparents
    info_sheet.cell(row = current_row + 1, column = 12).value = None
    
    #image order
    order_row = 2
    for i in trial_list:
        j = str(i)
        data_sheet.cell(row = order_row, column = 1).value = j
        order_row += 1
        
    #save file
    data_sheet.title = 'DATA' 
    excel.save('C:/Users/emily/Desktop/em_recog_static.xlsx')
    
def insert_data(excel, data_sheet, user_emo, img_num):
    data_sheet.cell(row = img_num, column = 3).value = user_emo.get()
    
    excel.save('C:/Users/emily/Desktop/em_recog_static.xlsx')

def collate(image_list_OG):
    name = userid.get()
    print(name, "Folder already exists, extracting information")
    user_file = open(name + '/' + 'User Information' + name + '.txt', 'r')
    info_list = []
    for i in user_file:
        info_list.append(i.strip('\n'))

    username = info_list[0]
    userage = info_list[1]
    usergender = info_list[2] #NEW
    userpsyc = info_list[4]
    userprof = info_list[5]
    useronset = info_list[6]
    userdegree = info_list[7]
    usersign = info_list[8]
    userlang = info_list[9]
    userparents = info_list[10]
    
    try:
        excel = load_workbook('C:/Users/emily/Desktop/em_recog_static_cleaned.xlsx')
        sort_sheet = excel['CLEANED']

        current_row = sort_sheet.max_row
    
        sort_sheet.cell(row = current_row + 1, column = 1).value = current_row 
        sort_sheet.cell(row = current_row + 1, column = 2).value = userage
        sort_sheet.cell(row = current_row + 1, column = 3).value = usergender #NEW
        sort_sheet.cell(row = current_row + 1, column = 4).value = useronset
        sort_sheet.cell(row = current_row + 1, column = 5).value = userdegree
        sort_sheet.cell(row = current_row + 1, column = 6).value = usersign
        sort_sheet.cell(row = current_row + 1, column = 7).value = userlang
        sort_sheet.cell(row = current_row + 1, column = 8).value = userparents

        excel.save(filename = 'C:/Users/emily/Desktop/em_recog_static_cleaned.xlsx')

    except FileNotFoundError:
        excel = Workbook()
        sort_sheet = excel.create_sheet("CLEANED", 0)

        data_column = 4
        while data_column < 9:
            letter_col = get_column_letter(data_column)
            sort_sheet.column_dimensions[letter_col].width = 20
            data_column += 1

        order_column = 9
        time_column = 69
        order_row = 1
        for i in image_list_OG:
            j = str(i)
            values = dictionary_img.getKey(j)
            letter = get_column_letter(order_column)
            letter_time = get_column_letter(time_column)
            sort_sheet.column_dimensions[letter].width = 20
            sort_sheet.column_dimensions[letter_time].width = 20
            sort_sheet.cell(row = order_row, column = order_column).value = j + " /" + values[0]
            sort_sheet.cell(row = order_row, column = time_column).value = j + " /" + values[0]
            order_column += 1
            time_column += 1

        sort_sheet.cell(row = 1, column = 2).value = "Age"
        sort_sheet.cell(row = 1, column = 3).value = "Gender" #NEW
        sort_sheet.cell(row = 1, column = 4).value = "Onset of Hearing Loss"
        sort_sheet.cell(row = 1, column = 5).value = "Severity"
        sort_sheet.cell(row = 1, column = 6).value = "Signing/Non-signing"
        sort_sheet.cell(row = 1, column = 7).value = "Language"
        sort_sheet.cell(row = 1, column = 8).value = "Parents"

        current_row = sort_sheet.max_row
    
        sort_sheet.cell(row = current_row + 1, column = 1).value = current_row 
        sort_sheet.cell(row = current_row + 1, column = 2).value = userage
        sort_sheet.cell(row = current_row + 1, column = 3).value = usergender #NEW
        sort_sheet.cell(row = current_row + 1, column = 4).value = useronset
        sort_sheet.cell(row = current_row + 1, column = 5).value = userdegree
        sort_sheet.cell(row = current_row + 1, column = 6).value = usersign
        sort_sheet.cell(row = current_row + 1, column = 7).value = userlang
        sort_sheet.cell(row = current_row + 1, column = 8).value = userparents

        excel.save(filename = 'C:/Users/emily/Desktop/em_recog_static_cleaned.xlsx')
        

def readData():
    studid = userid.get()
    file = open("C:/Users/emily/Desktop/PSYC15_Project_static2" + "/" + studid + "/" + "accuracy_static" + studid + ".txt", "r")
    listRead = file.readlines()
    acc = []
    count = 1

    for i in listRead:
        j = i.replace(str(count),"")
        k = j.replace(")", '')
        l = k.replace("\n", '')
        acc.append(l)
        count += 1
    file.close()
    
    file1 = open("C:/Users/emily/Desktop/PSYC15_Project_static2" + "/" + studid + "/" + "Time" + studid + ".txt", "r")
    listRead_time = file1.readlines()
    time = []

    del listRead_time[0]
    del listRead_time[0]
    for i in listRead_time:
        j = i.replace("\n", "")
        k = j[3:13]
        time.append(k)
    #print(time)
    file1.close()

            
    file2 = open("C:/Users/emily/Desktop/PSYC15_Project_static2" + "/" + studid + "/" + "Image Order" + studid + ".txt", "r")
    listRead1 = file2.readlines()
    
    img = []
    del listRead1[0]
    for i in listRead1:
        j = i.replace("\n", '')
        k = j[7:]
        img.append(int(k))

    file2.close()
    index = 0
    new_dict = {}
    time_dict = {}
    for i in range(0, len(img)):
        key = img[i]
        value = acc[index]
        time_value = time[index]
        index += 1
        new_dict[key] = value
        time_dict[key] = time_value
    

    excel = load_workbook('C:/Users/emily/Desktop/em_recog_static_cleaned.xlsx')
    sort_sheet = excel['CLEANED']

    current_row = sort_sheet.max_row
    order_column = 9
    time_column = 69

    x = sorted(new_dict)
    y = customSort(x) #this is the only time the algorithm is used

    timeX = sorted(time_dict)
    timeY = customSort(timeX)

    for i in y:
        j = new_dict[i]
        k = time_dict[i]
        sort_sheet.cell(row = current_row, column = order_column).value = j
        sort_sheet.cell(row = current_row, column = time_column).value = k
        order_column += 1
        time_column += 1

    excel.save(filename = 'C:/Users/emily/Desktop/em_recog_static_cleaned.xlsx')
    

def customSort(x): #The algorithm to give us a list index sorted by emotion
    count = 0
    index = 0
    position = 0
    iteration = 0
    newlist = []

    while iteration != 6:
        if count == 0:
            newlist.append(x[index])
            index += 6
            position += 1
            
            if position % 10 == 0:
                count += 1
                iteration += 1
                index = iteration
            else:
                count += 0
                iteration += 0
                
        if count == 1:
            newlist.append(x[index])
            index += 6
            position += 1

            if position % 10 == 0:
                count -= 1
                iteration += 1
                index = iteration
            else:
                count += 0
                iteration += 0

    return newlist


#recursive function to advance the imgs
def forward(img_num):
    global my_pic
    global start
    global submit
    global button_exit
    global studid
    global enter
    global trial_list
    global em_list
    start_w.title(' ')
    excel = load_workbook('C:/Users/emily/Desktop/em_recog_static.xlsx')
    data_sheet = excel['DATA']

#similar to windows
    my_pic.pack_forget()
    start.pack_forget()
    submit.pack_forget()
    enter.pack_forget()
    em_list.pack_forget()
    em_list.place_forget()

#trivial statement
    if img_num == 61:
        em_list.pack_forget()
        em_list.place_forget()
        my_pic = Label(start_w, image = imgLAST, bg = "white")
        my_pic.pack(side = "top", pady = (75,0))

        button_exit = Button(start_w, text = "Quit", command = lambda:[calcData.calcData(), response.organize(), goodbye(), collate(image_list_OG), readData()])
        
        button_exit.pack(side = "bottom", pady = (0, 300))

        goodbye_label = ttk.Label(start, text = "Please press quit.", background = "white")
        goodbye_label.pack(side = "bottom", pady = (0,0))

#advancing through imgs
    elif img_num == 1:
        my_pic = Label(start_w, image = trial_list[img_num - 1], bg = "white")
        my_pic.pack(side = "top", pady = (100,0))

        submit = Button(start_w, state = NORMAL, text = "Submit", command = lambda:[timeclick1(), appendclick(studid, click = user_emo.get()),\
                                                                 insert_data(excel, data_sheet, user_emo, img_num + 1),\
                                                                 mark_image(img_num - 1, trial_list, user_emo, studid, excel, data_sheet),\
                                                                 forward(img_num + 1), clear()])
        submit.bind("<Button-1>", buttoncall)
        submit.pack(side = "bottom", pady = (10, 200))

        enter = ttk.Entry(start_w, textvariable = user_emo)
        enter.bind("<Return>", buttoncall)
        enter.focus_set()
        enter.pack(side = 'bottom')

        em_list = Label(start_w, text = "HAPPY\n\nSAD\n\nANGRY\n\nSURPRISE\n\nDISGUST\n\nAFRAID", background = 'white')
        em_list.pack()
        em_list.place(x = 900, y = 500)
    
    elif img_num % 2 == 0:
        my_pic = Label(start_w, image = trial_list[img_num - 1], bg = "white")
        my_pic.pack(side = "top", pady = (100,0))

        submit = Button(start_w, text = "Submit", command = lambda:[timeclick(), appendclick(studid, click = user_emo.get()),\
                                                                 insert_data(excel, data_sheet, user_emo, img_num + 1),\
                                                                 mark_image(img_num - 1, trial_list, user_emo, studid, excel, data_sheet),\
                                                                 forward(img_num + 1), clear()])
        submit.bind("<Button-1>", buttoncall)
        submit.pack(side = "bottom", pady = (10, 200))

        enter = ttk.Entry(start_w, textvariable = user_emo)
        enter.bind("<Return>", buttoncall)
        enter.pack(side = 'bottom')

        em_list = Label(start_w, text = "HAPPY\n\nSAD\n\nANGRY\n\nSURPRISE\n\nDISGUST\n\nAFRAID", background = 'white')
        em_list.pack()
        em_list.place(x = 900, y = 500)
        
    elif img_num != 1 and img_num % 2 != 0 and img_num != 61:
        my_pic = Label(start_w, image = trial_list[img_num - 1], bg = "white")
        my_pic.pack(side = "top", pady = (100,0))
        
        submit = Button(start_w, text = "Submit", command = lambda:[timeclick1(), appendclick(studid, click = user_emo.get()),\
                                                                 insert_data(excel, data_sheet, user_emo, img_num + 1),\
                                                                 mark_image(img_num - 1, trial_list, user_emo, studid, excel, data_sheet),\
                                                                 forward(img_num + 1), clear()])
        submit.bind("<Button-1>", buttoncall)
        submit.pack(side = "bottom", pady = (10, 200))

        enter = ttk.Entry(start_w, textvariable = user_emo)
        enter.bind("<Return>", buttoncall)
        enter.pack(side = 'bottom')

        em_list = Label(start_w, text = "HAPPY\n\nSAD\n\nANGRY\n\nSURPRISE\n\nDISGUST\n\nAFRAID", background = 'white')
        em_list.pack()
        em_list.place(x = 900, y = 500)
        
#main
def windows():
#welcome page layout
    wel_label = ttk.Label(welcome_w, text = \
                          """\n\nWelcome!\n\n\n\nThe following page will ask you for the following information:
\nstudent ID
\nPlease proceed to the next page.""",\
                          background = 'white')
    wel_label.config(font = ("Calibri", 18), justify = 'center')
    wel_label.pack(side = "top", pady = (50, 0))

    #button layout
    next1 = ttk.Button(welcome_w, text = 'Next',\
                       command = lambda:[goto_info()])
    next1.bind("<Button-1>")
    next1.pack(side = "bottom", pady = (0, 250))

#info page layout
    info_label = ttk.Label(info_w, text = "Please enter your student ID number:", background = "white")
    info_label.config(font = ("Calibri", 12), justify = 'right')
    info_label.pack(side = "top", pady = (150, 0))

    #student id
    id_label = ttk.Label(info_w, text = "IVC or Saddleback Student ID number: ", background = 'white')
    id_label.pack()
    id_label.place(x = 494, y = 250)
    
    id_entry = ttk.Entry(info_w, textvariable = userid)
    id_entry.pack()
    id_entry.place(x = 700, y = 250)
    
    #advance
    confirm = ttk.Button(info_w, text = 'The above information is correct.',\
                         command = lambda:[folder(studid, trial_list), img_order(studid, trial_list), buttonpressed1()])
    confirm.bind("<Button-1>")
    confirm.pack()
    confirm.place(x = 590, y = 450)
                 
    next2 = ttk.Button(info_w, text = 'Next',\
                       command = lambda:[u_id(userid, studid), goto_instructions()])
    next2.bind("<Button-1>")

#instructions page layout
    instruct = ttk.Label(instruct_w, text = \
                   """Participation in this task is voluntary.\nYou may choose to opt out at any time and your data will not be recorded.
                    \nInstructions are below.
                    \nInstructions:\nFor each image, enter exactly one word to describe the emotion that you see as quickly and as accurately as possible,
                    then click the "Enter" key on the keyboard or the "Submit" button.
                    \nWhen you are done, please press the \"Quit\" button and leave the window up.
                    \nPlease ensure you have read all the instructions and ask the researcher if you have any questions.
                    \n\nPress the button below to confirm you have read the instructions and press \"Next\".""",\
                   background = 'white')
    instruct.config(font = ("Calibri", 18), justify = 'center')

    instruct.pack(side = "top", pady = (100, 0))

    confirm_dir_read = ttk.Button(instruct_w, text = 'I have read and I understand the above instructions.',\
                             command = lambda:[buttonpressed2()])
    confirm_dir_read.bind("<Button-1")
    confirm_dir_read.place(x = 545, y = 575)
    confirm_dir_read.pack()

    next3 = ttk.Button(instruct_w, text = 'Next',\
                       command = lambda:[goto_start()])
    next3.bind("<Button-1>") 
    
#declared variables for global/function use
#info
userid = StringVar()

#string variables
#user id
studid = userid.get()

#user entry
user_emo = StringVar()
#user_emo.trace("w", special)
enter = ttk.Entry(start_w, textvariable = user_emo)

#buttons
#info next
next2 = ttk.Button(info_w, text = 'Next',\
                       command = lambda:[goto_instructions()])

next3 = ttk.Button(instruct_w, text = 'Next',\
                       command = lambda:[goto_start()])

#start
start = ttk.Button(start_w, text = 'Begin',\
                   command = lambda:[forward(1)])
start.bind("<Button-1>", timeclickStart)
start.pack(side = 'bottom', pady = (0, 500))

#enter key
submit = Button(start_w, text = "Submit", command = lambda:[timeclick(), appendclick(studid, click = user_emo.get()),\
                                                                 insert_data(excel, data_sheet, user_emo, img_num + 1), forward(img_num + 1), clear()])

em_list = Label(start_w, text = "HAPPY\n\nSAD\n\nANGRY\n\nSURPRISE\n\nDISGUST\n\nAFRAID", background = 'white')

#quit
button_exit = Button(start_w, text = "Quit", command = lambda:[calcData.calcData(), response.organize(), goodbye()])

if __name__=="__main__":
    windows()
    
root.mainloop()
