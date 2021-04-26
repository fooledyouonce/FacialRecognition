#Emily Crowl
#Emotional Recognition: DYNAMIC 1
#PSYC15

#Collects user information:
#name, ID, PSYC class, PSYC professor
#onset of deficit, hearing ability, signing/nonsigning,
#and hearing ability of parents
#Runs 36 videos with each of the 6 basic emotions
#Users enter one word that describes which emotion they see
#program records response, marks right/wrong, and records time

from tkinter import ttk, Tk, Toplevel
from tkinter import *
import time
import random
import os
import os.path
import calcData
import response
import imageio
import threading
import moviepy.editor as mp
from openpyxl import *
from openpyxl.utils import get_column_letter
import dictionary_vid

#in case we are running this on a computer
#that doesn't have updated Python
try:
    from PIL import Image, ImageTk
except ModuleNotFoundError:
    import Image, ImageTk

#root window
root = Tk()

#initialize windows
#welcome window
welcome_w = Toplevel(root)
width_value = welcome_w.winfo_screenwidth()
height_value = welcome_w.winfo_screenheight()
welcome_w.configure(bg = 'white')
welcome_w.attributes('-fullscreen', True)
welcome_w.resizable(width = False, height = False)
welcome_w.title('Welcome')

#info windows
info1_w = Toplevel(root)
width_value = info1_w.winfo_screenwidth()
height_value = info1_w.winfo_screenheight()
info1_w.configure(bg = 'white')
info1_w.attributes('-fullscreen', True)
info1_w.resizable(width = False, height = False)
info1_w.title('User Information (1/2)')

info2_w = Toplevel(root)
width_value = info2_w.winfo_screenwidth()
height_value = info2_w.winfo_screenheight()
info2_w.configure(bg = 'white')
info2_w.attributes('-fullscreen', True)
info2_w.resizable(width = False, height = False)
info2_w.title('User Information (2/2)')

#informed consent window
informedc_w = Toplevel(root)
width_value = informedc_w.winfo_screenwidth()
height_value = informedc_w.winfo_screenheight()
informedc_w.configure(bg = 'white')
informedc_w.attributes('-fullscreen', True)
informedc_w.resizable(width = False, height = False)
informedc_w.title('Informed Consent')

#instructions window
instruct_w = Toplevel(root)
width_value = instruct_w.winfo_screenwidth()
height_value = instruct_w.winfo_screenheight()
instruct_w.configure(bg = 'white')
instruct_w.attributes('-fullscreen', True)
instruct_w.resizable(width = False, height = False)
instruct_w.title('Instructions')

#start window; manip in recursive function
start_w = Toplevel(root)
width_value = start_w.winfo_screenwidth()
height_value = start_w.winfo_screenheight()
start_w.configure(bg = 'white')
start_w.attributes('-fullscreen', True)
start_w.resizable(width = False, height = False)
start_w.title('Start')

#label we are manipulating in recursive function
my_pic = ttk.Label(start_w, text = "Press \"Begin\" when you are ready.",\
                   background = 'white')
my_pic.config(font = ("Calibri", 18), justify = 'center')

my_pic.pack(side = "top", pady = (100, 0))

###videos conversion
##img1 = "vid1.mp4"
##img2 = "vid2.mp4"
##img3 = "vid3.mp4"
##img4 = "vid4.mp4"
##img5 = "vid5.mp4"
##img6 = "vid6.mp4"
##img7 = "vid7.mp4"
##img8 = "vid8.mp4"
##img9 = "vid9.mp4"
##img10 = "vid10.mp4"
##img11 = "vid11.mp4"
##img12 = "vid12.mp4"
##img13 = "vid13.mp4"
##img14 = "vid14.mp4"
##img15 = "vid15.mp4"
##img16 = "vid16.mp4"
##img17 = "vid17.mp4"
##img18 = "vid18.mp4"
##img19 = "vid19.mp4"
##img20 = "vid20.mp4"
##img21 = "vid21.mp4"
##img22 = "vid22.mp4"
##img23 = "vid23.mp4"
##img24 = "vid24.mp4"
##img25 = "vid25.mp4"
##img26 = "vid26.mp4"
##img27 = "vid27.mp4"
##img28 = "vid28.mp4"
##img29 = "vid29.mp4"
##img30 = "vid30.mp4"
##img31 = "vid31.mp4"
##img32 = "vid32.mp4"
##img33 = "vid33.mp4"
##img34 = "vid34.mp4"
##img35 = "vid35.mp4"
##img36 = "vid36.mp4"
##
##imgLAST = PhotoImage(file = "LAST.png")
##
##image_list1 = [img1, img2, img3, img4, img5, img6, img7, img8,\
##               img9, img10, img11, img12, img13, img14, img15,\
##               img16, img17, img18, img19, img20, img21, img22,\
##               img23, img24, img25, img26, img27, img28, img29,\
##               img30, img31, img32, img33, img34, img35, img36]
##
###one time resize code, comment out when done
##image_list = []
##index = 0
##count = 111
##
##for i in image_list1:
##    
##    count1 = str(count)
##    clip = mp.VideoFileClip(image_list1[index])
##    clip_resized = clip.resize(height = 500)
##    clip_resized.write_videofile("vid" + count1 + ".mp4")
##    imgnew  = "vid" + count1 + ".mp4"
##    index += 1
##    count += 1
##    image_list.append(imgnew)

#converted videos
img1 = "vid111.mp4"
img2 = "vid112.mp4"
img3 = "vid113.mp4"
img4 = "vid114.mp4"
img5 = "vid115.mp4"
img6 = "vid116.mp4"
img7 = "vid117.mp4"
img8 = "vid118.mp4"
img9 = "vid119.mp4"
img10 = "vid120.mp4"
img11 = "vid121.mp4"
img12 = "vid122.mp4"
img13 = "vid123.mp4"
img14 = "vid124.mp4"
img15 = "vid125.mp4"
img16 = "vid126.mp4"
img17 = "vid127.mp4"
img18 = "vid128.mp4"
img19 = "vid129.mp4"
img20 = "vid130.mp4"
img21 = "vid131.mp4"
img22 = "vid132.mp4"
img23 = "vid133.mp4"
img24 = "vid134.mp4"
img25 = "vid135.mp4"
img26 = "vid136.mp4"
img27 = "vid137.mp4"
img28 = "vid138.mp4"
img29 = "vid139.mp4"
img30 = "vid140.mp4"
img31 = "vid141.mp4"
img32 = "vid142.mp4"
img33 = "vid143.mp4"
img34 = "vid144.mp4"
img35 = "vid145.mp4"
img36 = "vid146.mp4"

imgLAST = PhotoImage(file = "LAST.png")

image_list = [img1, img2, img3, img4, img5, img6, img7, img8,\
              img9, img10, img11, img12, img13, img14, img15,\
              img16, img17, img18, img19, img20, img21, img22,\
              img23, img24, img25, img26, img27, img28, img29,\
              img30, img31, img32, img33, img34, img35, img36]

image_list_OG = [img1, img2, img3, img4, img5, img6, img7, img8,\
              img9, img10, img11, img12, img13, img14, img15,\
              img16, img17, img18, img19, img20, img21, img22,\
              img23, img24, img25, img26, img27, img28, img29,\
              img30, img31, img32, img33, img34, img35, img36]
    
#image randomization
random.shuffle(image_list)
trial_list = image_list

#hides windows on command
root.withdraw()
info1_w.withdraw()
info2_w.withdraw()
informedc_w.withdraw()
instruct_w.withdraw
start_w.withdraw()

#functions to direct which window to open next
def goto_info1():
    welcome_w.destroy()
    info1_w.deiconify()
    
def goto_info2():
    info1_w.destroy()
    info2_w.deiconify()
    
def goto_informedc():
    info2_w.destroy()
    informedc_w.deiconify()

def goto_instructions():
    informedc_w.destroy()
    instruct_w.deiconify()

def goto_start():
    instruct_w.destroy()
    start_w.deiconify()

#functions for entries (Informed Consent and User Information txt files)
def folder(studid):
    name = userid.get()
    save_path = 'C:/Users/emily/Desktop/PSYC15_Project_dynamic1//' + name
    os.mkdir(save_path)

    save_path_D2 = 'C:/Users/emily/Desktop/PSYC15_Project_static2//' + name
    os.mkdir(save_path_D2)
    try:
        os.mkdir('C:/Users/emily/Desktop/PSYC15_Project_dynamic1//' + name)
        os.mkdir('C:/Users/emily/Desktop/PSYC15_Project_static2//' + name)
    except FileExistsError:
        print('C:/Users/emily/Desktop/PSYC15_Project_dynamic1//' + name, "Folder Created")
        print('C:/Users/emily/Desktop/PSYC15_Project_static2//' + name, "Folder Created")

#informed consent
def sig(usersig, studid):
    studid = userid.get()
    signature = usersig.get()
    save_path = 'C:/Users/emily/Desktop/PSYC15_Project_dynamic1//' + studid
    completeName = os.path.join(save_path, 'Informed Consent' + studid + '.txt')
    
    save_to_D2 = 'C:/Users/emily/Desktop/PSYC15_Project_static2//' + studid
    completeName_D2 = os.path.join(save_to_D2, 'Informed Consent' + studid + '.txt')
    outfile_D2 = open(completeName_D2, 'w')
    outfile_D2.write(str(signature) + '\n')
    outfile_D2.close()
    
    outfile = open(completeName, 'w')
    outfile.write(str(signature) + '\n')
    outfile.close()

#info1
def u_name(username, studid):
    studid = userid.get()
    name = username.get()
    save_path = 'C:/Users/emily/Desktop/PSYC15_Project_dynamic1//' + studid
    completeName = os.path.join(save_path, 'User Information' + studid + '.txt')

    save_to_D2 = 'C:/Users/emily/Desktop/PSYC15_Project_static2//' + studid
    completeName_D2 = os.path.join(save_to_D2, 'User Information' + studid + '.txt')
    outfile_D2 = open(completeName_D2, 'a')
    outfile_D2.write(str(name) + '\n')
    outfile_D2.close()

    outfile = open(completeName, 'a')
    outfile.write(str(name) + '\n')
    outfile.close()

def u_age(userage, studid):
    studid = userid.get()
    age = userage.get()
    save_path = 'C:/Users/emily/Desktop/PSYC15_Project_dynamic1//' + studid
    completeName = os.path.join(save_path, 'User Information' + studid + '.txt')

    save_to_D2 = 'C:/Users/emily/Desktop/PSYC15_Project_static2//' + studid
    completeName_D2 = os.path.join(save_to_D2, 'User Information' + studid + '.txt')
    outfile_D2 = open(completeName_D2, 'a')
    outfile_D2.write(str(age) + '\n')
    outfile_D2.close()

    outfile = open(completeName, 'a')
    outfile.write(str(age) + '\n')
    outfile.close()

def u_gender(userage, studid):
    studid = userid.get()
    gender = usergender.get()
    save_path = 'C:/Users/emily/Desktop/PSYC15_Project_dynamic1//' + studid
    completeName = os.path.join(save_path, 'User Information' + studid + '.txt')

    save_to_D2 = 'C:/Users/emily/Desktop/PSYC15_Project_static2//' + studid
    completeName_D2 = os.path.join(save_to_D2, 'User Information' + studid + '.txt')
    outfile_D2 = open(completeName_D2, 'a')
    outfile_D2.write(str(gender) + '\n')
    outfile_D2.close()

    outfile = open(completeName, 'a')
    outfile.write(str(gender) + '\n')
    outfile.close()

def u_id(userid, studid):
    studid = userid.get()
    save_path = 'C:/Users/emily/Desktop/PSYC15_Project_dynamic1//' + studid
    completeName = os.path.join(save_path, 'User Information' + studid + '.txt')

    save_to_D2 = 'C:/Users/emily/Desktop/PSYC15_Project_static2//' + studid
    completeName_D2 = os.path.join(save_to_D2, 'User Information' + studid + '.txt')
    outfile_D2 = open(completeName_D2, 'a')
    outfile_D2.write(str(studid) + '\n')
    outfile_D2.close()
    
    outfile = open(completeName, 'a')
    outfile.write(str(studid) + '\n')
    outfile.close()
    return studid

def u_psyc(userpsyc, studid):
    studid = userid.get()
    psyc = userpsyc.get()
    save_path = 'C:/Users/emily/Desktop/PSYC15_Project_dynamic1//' + studid
    completeName = os.path.join(save_path, 'User Information' + studid + '.txt')

    save_to_D2 = 'C:/Users/emily/Desktop/PSYC15_Project_static2//' + studid
    completeName_D2 = os.path.join(save_to_D2, 'User Information' + studid + '.txt')
    outfile_D2 = open(completeName_D2, 'a')
    outfile_D2.write(str(psyc) + '\n')
    outfile_D2.close()
    
    outfile = open(completeName, 'a')
    outfile.write(str(psyc) + '\n')
    outfile.close()

def u_prof(userprof, studid):
    studid = userid.get()
    prof = userprof.get()
    save_path = 'C:/Users/emily/Desktop/PSYC15_Project_dynamic1//' + studid
    completeName = os.path.join(save_path, 'User Information' + studid + '.txt')

    save_to_D2 = 'C:/Users/emily/Desktop/PSYC15_Project_static2//' + studid
    completeName_D2 = os.path.join(save_to_D2, 'User Information' + studid + '.txt')
    outfile_D2 = open(completeName_D2, 'a')
    outfile_D2.write(str(prof) + '\n')
    outfile_D2.close()
    
    outfile = open(completeName, 'a')
    outfile.write(str(prof) + '\n')
    outfile.close()

#info2
def u_onset(useronset, studid):
    studid = userid.get()
    onset = useronset.get()
    save_path = 'C:/Users/emily/Desktop/PSYC15_Project_dynamic1//' + studid
    completeName = os.path.join(save_path, 'User Information' + studid + '.txt')

    save_to_D2 = 'C:/Users/emily/Desktop/PSYC15_Project_static2//' + studid
    completeName_D2 = os.path.join(save_to_D2, 'User Information' + studid + '.txt')
    outfile_D2 = open(completeName_D2, 'a')
    outfile_D2.write(str(onset) + '\n')
    outfile_D2.close()
    
    outfile = open(completeName, 'a')
    outfile.write(str(onset) + '\n')
    outfile.close()
    
def u_degree(userdegree, studid):
    studid = userid.get()
    degree = userdegree.get()
    save_path = 'C:/Users/emily/Desktop/PSYC15_Project_dynamic1//' + studid
    completeName = os.path.join(save_path, 'User Information' + studid + '.txt')
    
    save_to_D2 = 'C:/Users/emily/Desktop/PSYC15_Project_static2//' + studid
    completeName_D2 = os.path.join(save_to_D2, 'User Information' + studid + '.txt')
    outfile_D2 = open(completeName_D2, 'a')
    outfile_D2.write(str(degree) + '\n')
    outfile_D2.close()
    
    outfile = open(completeName, 'a')
    outfile.write(str(degree) + '\n')
    outfile.close()

def u_sign(usersign, studid):
    studid = userid.get()
    sign = usersign.get()
    save_path = 'C:/Users/emily/Desktop/PSYC15_Project_dynamic1//' + studid
    completeName = os.path.join(save_path, 'User Information' + studid + '.txt')

    save_to_D2 = 'C:/Users/emily/Desktop/PSYC15_Project_static2//' + studid
    completeName_D2 = os.path.join(save_to_D2, 'User Information' + studid + '.txt')
    outfile_D2 = open(completeName_D2, 'a')
    outfile_D2.write(str(sign) + '\n')
    outfile_D2.close()
    
    outfile = open(completeName, 'a')
    outfile.write(str(sign) + '\n')
    outfile.close()

def u_lang(userlang, studid):
    studid = userid.get()
    lang = userlang.get()
    save_path = 'C:/Users/emily/Desktop/PSYC15_Project_dynamic1//' + studid
    completeName = os.path.join(save_path, 'User Information' + studid + '.txt')

    save_to_D2 = 'C:/Users/emily/Desktop/PSYC15_Project_static2//' + studid
    completeName_D2 = os.path.join(save_to_D2, 'User Information' + studid + '.txt')
    outfile_D2 = open(completeName_D2, 'a')
    outfile_D2.write(str(lang) + '\n')
    outfile_D2.close()
    
    outfile = open(completeName, 'a')
    outfile.write(str(lang) + '\n')
    outfile.close()

def u_parents(userparents, studid):
    studid = userid.get()
    parents = userparents.get()
    save_path = 'C:/Users/emily/Desktop/PSYC15_Project_dynamic1//' + studid
    completeName = os.path.join(save_path, 'User Information' + studid + '.txt')

    save_to_D2 = 'C:/Users/emily/Desktop/PSYC15_Project_static2//' + studid
    completeName_D2 = os.path.join(save_to_D2, 'User Information' + studid + '.txt')
    outfile_D2 = open(completeName_D2, 'a')
    outfile_D2.write(str(parents) + '\n')
    outfile_D2.close()
    
    outfile = open(completeName, 'a')
    outfile.write(str(parents) + '\n')
    outfile.close()

#text file for user accuracy
def accuracy(studid, mark, img_num):
    studid = userid.get()

    save_path = 'C:/Users/emily/Desktop/PSYC15_Project_dynamic1//' + studid
    completeName = os.path.join(save_path, 'accuracy' + studid + '.txt')
    
    save_to_D2 = 'C:/Users/emily/Desktop/PSYC15_Project_static2//' + studid
    completeName_D2 = os.path.join(save_to_D2, 'accuracy' + studid + '.txt')
    
    if not os.path.exists(save_path):
        outfile = open(completeName, 'w')
        outfile.write(str(img_num+1) + ') ' + str(mark) + '\n')
        outfile.close()
        
        outfile_D2 = open(completeName_D2, 'w')
        outfile_D2.write(str(img_num+1) + ') ' + str(mark) + '\n')
        outfile_D2.close()
        
    else:
        outfile = open(completeName, 'a')
        outfile.write(str(img_num+1) + ') ' + str(mark) + '\n')
        outfile.close()
        
        outfile_D2 = open(completeName_D2, 'a')
        outfile_D2.write(str(img_num+1) + ') ' + str(mark) + '\n')
        outfile_D2.close()

#text file for img order    
def img_order(studid, trial_list):
    studid = userid.get()
    save_path = 'C:/Users/emily/Desktop/PSYC15_Project_dynamic1//' + studid
    completeName = os.path.join(save_path, 'Image Order' + studid + '.txt')
    outfile = open(completeName, 'w')
    outfile.write(str('Random Image Order for ' + studid + '\n'))
    outfile.close
    for i in trial_list:
        save_path = 'C:/Users/emily/Desktop/PSYC15_Project_dynamic1//' + studid
        completeName = os.path.join(save_path, 'Image Order' + studid + '.txt')
        outfile = open(completeName, 'a')
        outfile.write(str(i) + '\n')
    outfile.close()

#accuracy checking
def mark_image(img_num, trial_list, user_emo, studid):
    if img_num != 37:
        user_answer = user_emo.get()
        
        video = str(trial_list[img_num])
        values = dictionary_vid.getKey(video)
        
        count = 0
        for i in values:
            if user_answer.lower() == i:
                count += 1

            else:
                count += 0
                
        if count != 1:
            mark = "I/V"
            accuracy(studid, mark, img_num)
            data_sheet.cell(row = img_num + 2, column = 4).value = mark
            data_sheet.cell(row = img_num + 2, column = 5).value = values[0]
            excel.save('C:/Users/emily/Desktop/em_recog_dynamic.xlsx')
            
        else:
            mark = "C"
            accuracy(studid, mark, img_num)
            data_sheet.cell(row = img_num + 2, column = 4).value = mark
            excel.save('C:/Users/emily/Desktop/em_recog_dynamic.xlsx')
    
#functions for buttons
def nextclick(studid):
    studid = userid.get()
    save_path = 'C:/Users/emily/Desktop/PSYC15_Project_dynamic1//' + studid
    completeName = os.path.join(save_path, 'User Information' + studid + '.txt')
    outfile = open(completeName, 'w')
    outfile.close()

def agreeclick(studid, click):
    studid = userid.get()
    save_path = 'C:/Users/emily/Desktop/PSYC15_Project_dynamic1//' + studid
    completeName = os.path.join(save_path, 'Informed Consent' + studid + '.txt')
    outfile = open(completeName, 'a')
    outfile.write(str(click) + '\n')
    outfile.close()

def writeclick(studid, click):
    studid = userid.get()
    click = user_emo.get()
    outfile = open('Raw Data.txt', 'a')
    outfile.write(str(click) + '\n')
    outfile.close()

def appendclick(studid, click):
    studid = userid.get()
    click = user_emo.get()
    outfile = open('Raw Data.txt', 'a') 
    outfile.write(str(click) + '\n')
    outfile.close()

#functions to write in times marked by button clicks
#function to capture clicks
def timeclickStart(event):
    studid = userid.get()
    start = time.time()
    outfile = open('Raw Data.txt', 'w')
    outfile.write(str(studid) + '\n')
    outfile.write(str(start) + '\n') 
    outfile.close()

def timeclick(*args):
    studid = userid.get()
    initial = time.time()
    outfile = open('Raw Data.txt', 'a')
    outfile.write(str(initial) + '\n')
    outfile.close()
    
def timeclick1(*args):
    studid = userid.get()
    post = time.time()
    outfile = open('Raw Data.txt', 'a')
    outfile.write(str(post) + '\n')
    outfile.close()

#function to invoke the submit button when enter key or mouse click is pressed
def buttoncall(*args):
    submit.invoke()

#functions for info next buttons
def buttonpressed1(*args):
    next2.pack(side = "bottom", pady = (0, 150))

def buttonpressed2(*args):
    next3.pack(side = "bottom", pady = (0, 150))

def buttonpressed3(*args):
    next4.pack(side = 'bottom', pady = (0, 100))

#pack submit and enter labels/clear entry box
def entrypack():
    submit.configure(state = NORMAL)
    submit.bind("<Button>", buttoncall)
    submit.pack(side = "bottom", pady = (5, 150))

    enter.configure(state = NORMAL)
    enter.bind("<Return>", buttoncall)
    enter.focus_set()
    enter.pack(side = 'bottom')

    em_list.pack()
    em_list.place(x = 900, y = 550)

#declare excel workbook globally
try:
    excel = load_workbook('C:/Users/emily/Desktop/em_recog_dynamic.xlsx')
    info_sheet = excel['INFO']
    data_sheet = excel.create_sheet("DATA", 1)
except FileNotFoundError:
    excel = Workbook()
    info_sheet = excel.create_sheet("INFO", 0)
    data_sheet = excel.create_sheet("DATA", 1)

def insert_info(username, userage, userid, userpsyc, userprof, useronset, userdegree, usersign, userlang, userparents, usersig):
    #info sheet
    info_sheet.column_dimensions['A'].width = 25
    info_sheet.column_dimensions['B'].width = 7
    info_sheet.column_dimensions['C'].width = 10
    info_sheet.column_dimensions['D'].width = 15
    info_sheet.column_dimensions['E'].width = 12
    info_sheet.column_dimensions['F'].width = 15
    info_sheet.column_dimensions['G'].width = 20
    info_sheet.column_dimensions['H'].width = 18
    info_sheet.column_dimensions['I'].width = 20
    info_sheet.column_dimensions['J'].width = 15
    info_sheet.column_dimensions['K'].width = 15
    info_sheet.column_dimensions['L'].width = 15

    #data sheet
    data_sheet.column_dimensions['A'].width = 15
    data_sheet.column_dimensions['B'].width = 15
    data_sheet.column_dimensions['C'].width = 15
    data_sheet.column_dimensions['D'].width = 20

    #info labels
    info_sheet.cell(row = 1, column = 1).value = "Name"
    info_sheet.cell(row = 1, column = 2).value = "Age"
    info_sheet.cell(row = 1, column = 3).value = "Gender"
    info_sheet.cell(row = 1, column = 4).value = "Student ID"
    info_sheet.cell(row = 1, column = 5).value = "Class"
    info_sheet.cell(row = 1, column = 6).value = "Professor"
    info_sheet.cell(row = 1, column = 7).value = "Onset hearing loss"
    info_sheet.cell(row = 1, column = 8).value = "Severity"
    info_sheet.cell(row = 1, column = 9).value = "Signing/non-signing"
    info_sheet.cell(row = 1, column = 10).value = "Language learning"
    info_sheet.cell(row = 1, column = 11).value = "Parents"
    info_sheet.cell(row = 1, column = 12).value = "IC"

    #data labels
    data_sheet.cell(row = 1, column = 1).value = "Image Order"
    data_sheet.cell(row = 1, column = 2).value = "Time"
    data_sheet.cell(row = 1, column = 3).value = "Response"
    data_sheet.cell(row = 1, column = 4).value = "Correct/Incorrect"
    data_sheet.cell(row = 1, column = 5).value = "Answer"    

    #info 
    #get method returns current text as string, placing in specific location
    current_row = info_sheet.max_row #THE ONE LINE OF CODE THAT SEPARATED GOOD AND EVIL
    
    info_sheet.cell(row = current_row + 1, column = 1).value = username.get()
    info_sheet.cell(row = current_row + 1, column = 2).value = userage.get()
    info_sheet.cell(row = current_row + 1, column = 3).value = usergender.get()
    info_sheet.cell(row = current_row + 1, column = 4).value = userid.get()
    info_sheet.cell(row = current_row + 1, column = 5).value = userpsyc.get()
    info_sheet.cell(row = current_row + 1, column = 6).value = userprof.get()
    info_sheet.cell(row = current_row + 1, column = 7).value = useronset.get()
    info_sheet.cell(row = current_row + 1, column = 8).value = userdegree.get()
    info_sheet.cell(row = current_row + 1, column = 9).value = usersign.get()
    info_sheet.cell(row = current_row + 1, column = 10).value = userlang.get()
    info_sheet.cell(row = current_row + 1, column = 11).value = userparents.get()
    info_sheet.cell(row = current_row + 1, column = 12).value = usersig.get()
    
    #image order
    order_row = 2
    for i in trial_list:
        j = str(i)
        data_sheet.cell(row = order_row, column = 1).value = j
        order_row += 1
        
    #save file
    data_sheet.title = userid.get()   
    excel.save('C:/Users/emily/Desktop/em_recog_dynamic.xlsx')
    
def insert_data(userid, user_emo, img_num):
    data_sheet.cell(row = img_num, column = 3).value = user_emo.get()
    excel.save('C:/Users/emily/Desktop/em_recog_dynamic.xlsx')

def collate(image_list_OG):
    try:
        excel = load_workbook('C:/Users/emily/Desktop/em_recog_dynamic_cleaned.xlsx')
        sort_sheet = excel['CLEANED']

        current_row = sort_sheet.max_row
    
        sort_sheet.cell(row = current_row + 1, column = 1).value = current_row 
        sort_sheet.cell(row = current_row + 1, column = 2).value = userage.get()
        sort_sheet.cell(row = current_row + 1, column = 3).value = usergender.get()
        sort_sheet.cell(row = current_row + 1, column = 4).value = useronset.get()
        sort_sheet.cell(row = current_row + 1, column = 5).value = userdegree.get()
        sort_sheet.cell(row = current_row + 1, column = 6).value = usersign.get()
        sort_sheet.cell(row = current_row + 1, column = 7).value = userlang.get()
        sort_sheet.cell(row = current_row + 1, column = 8).value = userparents.get()

        excel.save(filename = 'C:/Users/emily/Desktop/em_recog_dynamic_cleaned.xlsx')

    except FileNotFoundError:
        excel = Workbook()
        sort_sheet = excel.create_sheet("CLEANED", 0)

        data_column = 4
        while data_column < 9:
            letter_col = get_column_letter(data_column)
            sort_sheet.column_dimensions[letter_col].width = 20
            data_column += 1

        order_column = 9
        time_column = 45
        order_row = 1
        for i in image_list_OG:
            j = str(i)
            values = dictionary_vid.getKey(j)
            letter = get_column_letter(order_column)
            letter_time = get_column_letter(time_column)
            sort_sheet.column_dimensions[letter].width = 20
            sort_sheet.column_dimensions[letter_time].width = 20
            sort_sheet.cell(row = order_row, column = order_column).value = j + " /" + values[0]
            sort_sheet.cell(row = order_row, column = time_column).value = j + " /" + values[0]
            order_column += 1
            time_column += 1

        sort_sheet.cell(row = 1, column = 2).value = "Age"
        sort_sheet.cell(row = 1, column = 3).value = "Gender"
        sort_sheet.cell(row = 1, column = 4).value = "Onset of Hearing Loss"
        sort_sheet.cell(row = 1, column = 5).value = "Severity"
        sort_sheet.cell(row = 1, column = 6).value = "Signing/Non-signing"
        sort_sheet.cell(row = 1, column = 7).value = "Language"
        sort_sheet.cell(row = 1, column = 8).value = "Parents"

        current_row = sort_sheet.max_row
    
        sort_sheet.cell(row = current_row + 1, column = 1).value = current_row 
        sort_sheet.cell(row = current_row + 1, column = 2).value = userage.get()
        sort_sheet.cell(row = current_row + 1, column = 3).value = usergender.get()
        sort_sheet.cell(row = current_row + 1, column = 4).value = useronset.get()
        sort_sheet.cell(row = current_row + 1, column = 5).value = userdegree.get()
        sort_sheet.cell(row = current_row + 1, column = 6).value = usersign.get()
        sort_sheet.cell(row = current_row + 1, column = 7).value = userlang.get()
        sort_sheet.cell(row = current_row + 1, column = 8).value = userparents.get()

        excel.save(filename = 'C:/Users/emily/Desktop/em_recog_dynamic_cleaned.xlsx')
        
def readData():
    studid = userid.get()
    file = open("C:/Users/emily/Desktop/PSYC15_Project_dynamic1//" + "/" + studid + "/" + "accuracy" + studid + ".txt", "r")
    listRead = file.readlines()
    acc = []
    count = 1

    for i in listRead:
        j = i.replace(str(count),"")
        k = j.replace(")", '')
        l = k.replace("\n", '')
        acc.append(l)
        count += 1
    file.close()

    file2 = open("C:/Users/emily/Desktop/PSYC15_Project_dynamic1" + "/" + studid + "/" + "Time" + studid + ".txt", "r")
    listRead_time = file2.readlines()
    time = []

    del listRead_time[0]
    del listRead_time[0]
    for i in listRead_time:
        j = i.replace("\n", "")
        k = j[3:13]
        time.append(k)
    file2.close()
    
    studid = userid.get()
    file3 = open("C:/Users/emily/Desktop/PSYC15_Project_dynamic1//" + "/" + studid + "/" + "Image Order" + studid + ".txt", "r")
    listRead1 = file3.readlines()
    
    img = []
    del listRead1[0]
    for i in listRead1:
        j = i.replace("\n", '')
        k = j[3:6]
        img.append(int(k))
    file3.close() 
    
    index = 0
    time_dict= {}
    new_dict = {}
    for i in range(0, len(img)):
        key = img[i]
        value = acc[index]
        time_value = time[index]
        index += 1
        new_dict[key] = value
        time_dict[key] = time_value

    excel = load_workbook('C:/Users/emily/Desktop/em_recog_dynamic_cleaned.xlsx')
    sort_sheet = excel['CLEANED']

    current_row = sort_sheet.max_row
    order_column = 9
    time_column = 45

    x = sorted(new_dict)
    y = sorted(x)

    for i in y:
        j = new_dict[i]
        k = time_dict[i]
        sort_sheet.cell(row = current_row, column = order_column).value = j
        sort_sheet.cell(row = current_row, column = time_column).value = k
        order_column += 1
        time_column += 1
    
    excel.save(filename = 'C:/Users/emily/Desktop/em_recog_dynamic_cleaned.xlsx')
    
#function for goodbye window
def goodbye():
    root.title("Thank you")
    global button_exit
    button_exit.pack_forget()

    end_label = ttk.Label(start_w, text = """\n\n\nThank you for your participation in this task.
                                         \nPlease wait as the next task is being prepared for you.""", \
                          background = "white")
    end_label.config(font = ("Calibri", 15), justify = 'center')
    
    end_label.pack(side = "top", pady = (20, 0))

#recursive function to advance the imgs
def forward(img_num):
    global my_pic
    global start
    global submit
    global button_exit
    global studid
    global enter
    global em_list
    start_w.title(' ')

#similar to windows
    my_pic.pack_forget()
    start.pack_forget()
    submit.pack_forget()
    enter.pack_forget()
    em_list.pack_forget()
    em_list.place_forget()

    enter.delete(0, END)
    submit.configure(state = DISABLED)
    enter.configure(state = DISABLED)

#trivial statement
    if img_num == 37:
        em_list.pack_forget()
        em_list.place_forget()
        my_pic = Label(start_w, image = imgLAST, bg = "white")
        my_pic.pack(side = "top", pady = (75,0))

        button_exit = Button(start_w, text = "Quit", command = lambda:[calcData.calcData(), response.organize(), goodbye(),\
                                                                       collate(image_list_OG), readData()])
        
        button_exit.pack(side = "bottom", pady = (0, 300))

        goodbye_label = ttk.Label(start, text = "Please press quit.", background = "white")
        goodbye_label.pack(side = "bottom", pady = (0,0))

#advancing through imgs
    elif img_num == 1:
        def stream(label):
            video = imageio.get_reader(trial_list[img_num - 1])
            
            for vid in video.iter_data():
                frame_vid = ImageTk.PhotoImage(Image.fromarray(vid))
                label.config(image = frame_vid)
                label.image = frame_vid

        thread = threading.Thread(target=stream, args = (my_pic,))
        thread.start() 

        my_pic.pack(side = "top")

        submit = Button(start_w, text = "Submit", state = DISABLED, command = lambda:[timeclick1(), appendclick(studid, click = user_emo.get()),\
                                                                insert_data(userid, user_emo, img_num + 1),\
                                                                mark_image(img_num - 1, trial_list, user_emo, studid),\
                                                                forward(img_num + 1)])

        em_list = Label(start_w, text = "HAPPY\n\nSAD\n\nANGRY\n\nSURPRISE\n\nDISGUST\n\nAFRAID", background = 'white')

        enter = ttk.Entry(start_w, state = DISABLED, textvariable = user_emo)

        submit.after(5500, entrypack)
    
    elif img_num % 2 == 0:
        def stream(label):
            video = imageio.get_reader(trial_list[img_num - 1])
            
            for vid in video.iter_data():
                frame_vid = ImageTk.PhotoImage(Image.fromarray(vid))
                label.config(image = frame_vid)
                label.image = frame_vid

        thread = threading.Thread(target=stream, args = (my_pic,))
        thread.start()

        my_pic.pack(side = "top")

        submit = Button(start_w, text = "Submit", state = DISABLED, command = lambda:[timeclick(), appendclick(studid, click = user_emo.get()),\
                                                                insert_data(userid, user_emo, img_num + 1),\
                                                                mark_image(img_num - 1, trial_list, user_emo, studid),\
                                                                forward(img_num + 1)])
        
        em_list = Label(start_w, text = "HAPPY\n\nSAD\n\nANGRY\n\nSURPRISE\n\nDISGUST\n\nAFRAID", background = 'white')

        enter = ttk.Entry(start_w, textvariable = user_emo)

        submit.after(5500, entrypack)
        
    elif img_num != 1 and img_num % 2 != 0 and img_num != 37:
        def stream(label):
            video = imageio.get_reader(trial_list[img_num - 1])
            
            for vid in video.iter_data():
                frame_vid = ImageTk.PhotoImage(Image.fromarray(vid))
                label.config(image = frame_vid)
                label.image = frame_vid

        thread = threading.Thread(target = stream, args = (my_pic,))
        thread.start()

        my_pic.pack(side = "top")
        
        submit = Button(start_w, text = "Submit", state = DISABLED, command = lambda:[timeclick1(), appendclick(studid, click = user_emo.get()),\
                                                                insert_data(userid, user_emo, img_num + 1),\
                                                                mark_image(img_num - 1, trial_list, user_emo, studid),\
                                                                forward(img_num + 1)])

        em_list = Label(start_w, text = "HAPPY\n\nSAD\n\nANGRY\n\nSURPRISE\n\nDISGUST\n\nAFRAID", background = 'white')

        enter = ttk.Entry(start_w, textvariable = user_emo)
        
        submit.after(5500, entrypack)
        
#main
def windows():
#welcome page layout
    wel_label = ttk.Label(welcome_w, text = \
                          """\n\nWelcome!\n\nThe following pages will ask you for the following information:
\nname, student ID, psychology class, psychology professor,
onset of hearing loss, degree of severity of hearing loss,
age of language acquisition, and hearing ability of parents
\n\nYou will also read and digitally sign an informed consent form.
\nPlease proceed to the next page.""",\
                          background = 'white')
    wel_label.config(font = ("Calibri", 18), justify = 'center')
    wel_label.pack(side = "top", pady = (50, 0))

    #button layout
    next1 = ttk.Button(welcome_w, text = 'Next',\
                       command = lambda:[goto_info1()])
    next1.bind("<Button-1>")
    next1.pack(side = "bottom", pady = (0, 250))

#info1 page layout
    info1_label = ttk.Label(info1_w, text = "Please enter all of the following information below:", background = "white")
    info1_label.config(font = ("Calibri", 12), justify = 'right')
    info1_label.pack(side = "top", pady = (150, 0))

    #name
    name_label = ttk.Label(info1_w, text = "Name (first and last): ", background = 'white')
    name_label.pack()
    name_label.place(x = 540, y = 200)
    
    name_entry = ttk.Entry(info1_w, textvariable = username)
    name_entry.pack()
    name_entry.place(x = 670, y = 200)

    #age
    age_label = ttk.Label(info1_w, text = "Age: ", background = 'white')
    age_label.pack()
    age_label.place(x = 625, y = 250)
    
    age_entry = ttk.Entry(info1_w, textvariable = userage)
    age_entry.pack()
    age_entry.place(x = 670, y = 250)

    #gender
    gender_label = ttk.Label(info1_w, text = "Gender: ", background = 'white')
    gender_label.pack()
    gender_label.place(x = 610, y = 300)
    
    gender_entry = ttk.Entry(info1_w, textvariable = usergender)
    gender_entry.pack()
    gender_entry.place(x = 670, y = 300)

    #student id
    id_label = ttk.Label(info1_w, text = "IVC or Saddleback student ID number: ", background = 'white')
    id_label.pack()
    id_label.place(x = 450, y = 350)
    
    id_entry = ttk.Entry(info1_w, textvariable = userid)
    id_entry.pack()
    id_entry.place(x = 670, y = 350)

    #psyc class
    psyc_label = ttk.Label(info1_w, text = "Psychology class you are currently enrolled in: ", background = 'white')
    psyc_label.pack()
    psyc_label.place(x = 405, y = 400)
    
    psyc_entry = ttk.Entry(info1_w, textvariable = userpsyc)
    psyc_entry.pack()
    psyc_entry.place(x = 670, y = 400)

    #professor name
    prof_label = ttk.Label(info1_w, text = "Psychology professor: ", background = 'white')
    prof_label.pack()
    prof_label.place(x = 533, y = 450)
    
    prof_entry = ttk.Entry(info1_w, textvariable = userprof)
    prof_entry.pack()
    prof_entry.place(x = 670, y = 450)
    
    #advance
    confirm = ttk.Button(info1_w, text = 'The above information is correct.',\
                         command = lambda:[folder(studid), nextclick(studid), buttonpressed1()])
    confirm.bind("<Button-1>")
    confirm.pack()
    confirm.place(x = 590, y = 550)
                 
    next2 = ttk.Button(info1_w, text = 'Next',\
                       command = lambda:[u_name(username, studid), u_age(userage, studid), u_gender(userage, studid), u_id(userid, studid), u_psyc(userpsyc, studid), u_prof(userprof, studid), goto_info2()])
    next2.bind("<Button-1>")

#info2 page layout
    info2_label = ttk.Label(info2_w, text = "Please enter all of the following information below:", background = "white")
    info2_label.config(font = ("Calibri", 12), justify = 'center')
    info2_label.pack(side = "top", pady = (150, 0))

    #onset of hearing loss
    onset_label = ttk.Label(info2_w, text = "Onset of hearing loss (if hearing, please enter N/A): ", background = 'white')
    onset_label.config(justify = 'center')
    onset_label.pack()
    onset_label.place(x = 555, y = 200)
    
    onset_entry = ttk.Entry(info2_w, textvariable = useronset)
    onset_entry.pack()
    onset_entry.place(x = 635, y = 225)

    #degree of severity
    degree_label = ttk.Label(info2_w, text = """Severity of hearing loss (deaf = profound hearing loss,
hard-of-hearing = medium hearing loss, hearing = little-to-no hearing loss: """, background = 'white')
    degree_label.pack()
    degree_label.place(x = 555, y = 265)
    
    degree_entry = ttk.Entry(info2_w, textvariable = userdegree)
    degree_entry.pack()
    degree_entry.place(x = 635, y = 300)

    #signing
    sign_label = ttk.Label(info2_w, text = "Are you fluent in sign language? (Y/N): ", background = 'white')
    sign_label.pack()
    sign_label.place(x = 555, y = 340)
    
    sign_entry = ttk.Entry(info2_w, textvariable = usersign)
    sign_entry.pack()
    sign_entry.place(x = 635, y = 370)

    #age of language acquisition
    lang_label = ttk.Label(info2_w, text = "Was your language learning typical, delayed, or advanced?: ", background = 'white')
    lang_label.pack()
    lang_label.place(x = 555, y = 410)
    
    lang_entry = ttk.Entry(info2_w, textvariable = userlang)
    lang_entry.pack()
    lang_entry.place(x = 635, y = 440)

    #parents hearing ability
    parents_label = ttk.Label(info2_w, text = "Hearing ability of parents (hearing/nonhearing): ", background = 'white')
    parents_label.pack()
    parents_label.place(x = 555, y = 470)
    
    parents_entry = ttk.Entry(info2_w, textvariable = userparents)
    parents_entry.pack()
    parents_entry.place(x = 635, y = 500)

    #advance
    confirm2 = ttk.Button(info2_w, text = 'The above information is correct.',\
                         command = lambda:[buttonpressed2()])
    confirm2.bind("<Button-1>")
    confirm2.pack()
    confirm2.place(x = 590, y = 550)
    
    next3 = ttk.Button(info2_w, text = 'Next',\
                       command = lambda:[u_onset(useronset, studid), u_degree(userdegree, studid), u_sign(usersign, studid),\
                                         u_lang(userlang, studid), u_parents(userparents, studid), goto_informedc()])
    next3.bind("<Button-1>") 

#informed consent page layout
    ic_label = ttk.Label(informedc_w, text = """\nThis is an informed consent form for the research conducted by Emily Crowl at Irvine Valley College.
                                            \nResearcher: Emily Crowl
                                            \nAdvisor: Benjamin Mis
                                            \nInstitution: Irvine Valley College
                                            \nPurpose:
This research is being conducted to further the understanding of visual processing across different levels of hearing.

                                            \nIntroduction:
The visual processing in individuals is being investigated, after it was discovered that hearing potentially influences
other abilities that are taken for granted.  In doing background research, it was found that there was data on deaf and
hearing individuals, but little-to-none on hard-of-hearing individuals.  It is hoped that the results of this study will
give insight into how hard-of-hearing individuals process the world.

                                            \nParticipation Information:
It will take approximately 20 minutes to complete the study.
You will be asked to complete (2) tasks, each involving you to respond to an image or a video.
Further instructions are on the next window.

                                            \nConsent to Participate:
Your participation in this research is fully voluntary and all of your personal information will remain anonymous.
You may discontinue the study at any time without any penalty or loss of benefits to which you are entitled.
Refusal to participate will result in no penalty or loss of benefits to which you are entitled. 
If you have any questions before you begin the task, do not hesitate to ask the researcher.
There are no potential risks from participating in this research.
You will receive credit in your psychology class for participating in this research.

If you have further questions about the study, or would like your data removed please email the researcher at:
\nemilycrowl2k@gmail.com

Contact Loris Fargoli, Director, Office of Research, Planning and Accreditation, at 949-451-5513 for answers
to pertinent questions about the research and research participants’ rights, and in the event of a research-related injury.

I have read the foregoing information, or it has been read to me.
I have had the opportunity to ask questions about it and any questions that I have asked have been answered to my satisfaction.
I consent to voluntarily participate as a participant in this research.
Please enter your name and press "I agree".""",\
                          background = 'white')
    ic_label.config(justify = 'center')
    ic_label.pack(side = "top", pady = (0, 20))

    sig_entry = ttk.Entry(informedc_w, textvariable = usersig)
    sig_entry.pack()
    sig_entry.place(x = 625, y = 690)
    
    #button layout
    agree = ttk.Button(informedc_w, text = 'I agree',\
                       command = lambda:[img_order(studid, trial_list), sig(usersig, studid), agreeclick(studid, click = 'agree'), goto_instructions()\
                                        ,insert_info(username, userage, userid, userpsyc, userprof, useronset, userdegree, usersign, userlang, userparents, usersig)])
    agree.bind("<Button-1>")
    agree.pack()
    agree.place(x = 650, y = 715)

#instructions page layout
    instruct = ttk.Label(instruct_w, text = \
                   """Participation in this task is voluntary.\nYou may choose to opt out at any time and your data will not be recorded.
                    \nInstructions are below.
                    \nInstructions:\nFor each video, enter a word to describe the emotion that you see as quickly and as accurately as possible,
                    then click the "Enter" key on the keyboard or the "Submit" button.
                    \n***Let the video play in its entirety***
                    \nWhen you are done, please press the \"Quit\" button and leave the window up.
                    \nPlease ensure you have read all the instructions and ask the researcher if you have any questions.
                    \n\nPress the button below to confirm you have read the instructions and press \"Next\".""",\
                   background = 'white')
    instruct.config(font = ("Calibri", 18), justify = 'center')

    instruct.pack(side = "top", pady = (100, 0))

    confirm_dir_read = ttk.Button(instruct_w, text = 'I have read and I understand the above instructions.',\
                             command = lambda:[buttonpressed3()])
    confirm_dir_read.bind("<Button-1")
    confirm_dir_read.place(x = 545, y = 575)
    confirm_dir_read.pack()

    next4 = ttk.Button(instruct_w, text = 'Next',\
                       command = lambda:[goto_start()])
    next4.bind("<Button-1>")

#declared variables for global/function use
#info1
usersig = StringVar()
username = StringVar()
userage = StringVar()
usergender = StringVar()
userid = StringVar()
userpsyc = StringVar()
userprof = StringVar()

#string variables
#user id
studid = userid.get()

#info2
useronset = StringVar()
userdegree = StringVar()
usersign = StringVar()
userlang = StringVar()
userparents = StringVar()

#user entry
user_emo = StringVar()

enter = ttk.Entry(start_w, textvariable = user_emo)

#buttons
#info next
next2 = ttk.Button(info1_w, text = 'Next',\
                       command = lambda:[u_name(username, studid), u_age(userage, studid), u_gender(usergender, studid), u_id(userid, studid), u_psyc(userpsyc, studid), u_prof(userprof, studid), goto_info2()])

next3 = ttk.Button(info2_w, text = 'Next',\
                       command = lambda:[u_onset(useronset, studid), u_degree(userdegree, studid), u_sign(usersign, studid),\
                                         u_lang(userlang, studid), u_parents(userparents, studid), goto_informedc()])

next4 = ttk.Button(instruct_w, text = 'Next',\
                       command = lambda:[goto_start()])

#start
start = ttk.Button(start_w, text = 'Begin',\
                       command = lambda:[forward(1)])
start.bind("<Button-1>", timeclickStart)
start.pack(side = 'bottom', pady = (0, 500))

#enter key
submit = Button(start_w, text = "Submit", command = lambda:[timeclick(), appendclick(studid, click = user_emo.get()), forward(1)])

em_list = Label(start_w, text = "HAPPY\n\nSAD\n\nANGRY\n\nSURPRISE\n\nDISGUST\n\nAFRAID", background = 'white')

#quit
button_exit = Button(start_w, text = "Quit", command = lambda:[calcData.calcData(), response.organize(), goodbye()])

if __name__=="__main__":
    windows()
    
root.mainloop()
